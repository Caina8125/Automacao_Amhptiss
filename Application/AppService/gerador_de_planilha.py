import pandas as pd
from tkinter import filedialog
import tkinter.messagebox
from openpyxl import load_workbook
from datetime import datetime

def substituir_pontos(valor):   
    return valor.replace('.0', '')

def fatura_atualizada(numero_autorizacao, numero_operadora, lista_autorizacao):
    if (numero_autorizacao in lista_autorizacao) or (numero_operadora in lista_autorizacao):
        return True
    else:
        return False

def verificar_frame_vazio(df_filtrado, df, coluna, numero):
    if df_filtrado.empty == True:
        fatura_df = df.loc[(df[coluna] == numero)]
        return fatura_df
    else:
        return df_filtrado
    
def is_datetime(obj):
    return isinstance(obj, datetime)

def gerar_planilha():
    try:
        tkinter.messagebox.showinfo('Planilha Detalhado Unificado', 'Selecione a planilha do relatório Detalhado Normal/Especial (Unificado).')
        planilha_unificada = filedialog.askopenfilename()
        tkinter.messagebox.showinfo('Planilha Demonstrativo', 'Selecione a planilha do demonstrativo do GDF')
        planilha_gdf = filedialog.askopenfilename()
        df_plan_unificada = pd.read_excel(planilha_unificada)
        # df_plan_unificada['AUTORIZACAO'] = df_plan_unificada['AUTORIZACAO'].astype(str)
        df_plan_gdf = pd.read_excel(planilha_gdf)
        df_plan_gdf['Autorização'] = df_plan_gdf['Autorização'].astype(str)
        df_plan_gdf['Autorização'] = df_plan_gdf['Autorização'].apply(substituir_pontos)
        lista_autorizacao_nova = df_plan_gdf['Autorização'].values.tolist()
        lista = []
        count_nao_encontradas = 0

        for i, l in df_plan_unificada.iterrows():
            senha_plan_uni = str(l['AUTORIZACAO']).replace('.0', '')
            numero_op_plan_uni = str(l['GUIAATENDIMENTO']).replace('.0', '')
            fatura_recurso_plan_uni = str(l['PROCESSOID']).replace('.0', '')
            realizado_plan_uni = l['DATAREALIZADO']

            if is_datetime(realizado_plan_uni):
                realizado_plan_uni = realizado_plan_uni.strftime('%d/%m/%Y')

            procedimento_plan_uni = str(l['CODIGOID']).replace('.0', '')
            controle_plan_uni = str(l['ATENDIMENTOID']).replace('.0', '')
            if senha_plan_uni.isdigit():
                fatura_df_gdf_filtrado = df_plan_gdf.loc[(df_plan_gdf['Autorização Origem'] == int(senha_plan_uni))]
                fatura_df_gdf = verificar_frame_vazio(fatura_df_gdf_filtrado, df_plan_gdf, 'Autorização Origem', senha_plan_uni)
            else:
                fatura_df_gdf_filtrado = df_plan_gdf.loc[(df_plan_gdf['Autorização Origem'] == senha_plan_uni)]

            if fatura_atualizada(senha_plan_uni, numero_op_plan_uni, lista_autorizacao_nova):
                count_nao_encontradas += 1
                dados = {"Encontrada" : ['Esta autorização já está atualizada']}
                df_dados = pd.DataFrame(dados)
                book = load_workbook(planilha_unificada)
                writer = pd.ExcelWriter(planilha_unificada, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df_dados.to_excel(writer, startrow= i + 1, startcol=39, header=False, index=False)
                writer.save()
                continue

            if fatura_df_gdf.empty == True:
                count_nao_encontradas += 1
                dados = {"Encontrada" : ['O número da autorização e o número da operadora desta guia não foram encontradas no demonstrativo']}
                df_dados = pd.DataFrame(dados)
                book = load_workbook(planilha_unificada)
                writer = pd.ExcelWriter(planilha_unificada, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df_dados.to_excel(writer, sheet_name='Cobranca_Unificada', startrow= i + 1, startcol=39, header=False, index=False)
                writer.save()
                continue

            for ind, lin in fatura_df_gdf.iterrows():
                numero_senha_plan_gdf = str(lin['Autorização Origem']).replace('.0', '')
                data_de_atendimento_gdf = lin['Data de Realização']
                codigo_gdf = str(lin['Código']).replace('.0', '')
                autorizacao_nova = str(lin['Autorização']).replace('.0', '')

                comparacao_senha = senha_plan_uni == numero_senha_plan_gdf
                comparacao_numero = numero_op_plan_uni == numero_senha_plan_gdf
                comparacao_data = realizado_plan_uni == data_de_atendimento_gdf
                comparacao_codigo = procedimento_plan_uni == codigo_gdf

                if (comparacao_senha or comparacao_numero) and comparacao_data and comparacao_codigo:
                    lista_linha = [controle_plan_uni, autorizacao_nova, senha_plan_uni, numero_op_plan_uni, '', fatura_recurso_plan_uni]
                    break

            lista.append(lista_linha)

        if len(lista) > 0:
            cabecalho = ['Controle', 'Autorização Nova', 'Autorização Original', 'Nro. Guia', 'Fatura Inicial', 'Fatura Recurso']
            df_nova_planilha = pd.DataFrame(lista)
            df_nova_planilha.columns = cabecalho
            data_e_hora_atuais = datetime.now()
            data_e_hora_em_texto = data_e_hora_atuais.strftime('%d_%m_%Y_%H_%M')
            segundo = data_e_hora_atuais.second
            df_nova_planilha.to_excel(f'GDF\\GDF_{data_e_hora_em_texto}_{segundo}.xlsx', index=False)
            if count_nao_encontradas > 0:
                tkinter.messagebox.showinfo("Gerador de Planilha", f"Planilha gerada!\nTotal de linhas não encontradas: {count_nao_encontradas}")
            else:
                tkinter.messagebox.showinfo("Gerador de Planilha", f"Planilha gerada com sucesso!")

        else:
            tkinter.messagebox.showinfo("Gerador de Planilha", f"A Planilha não foi gerada!\nTotal de linhas não encontradas: {count_nao_encontradas}")
    
    except Exception as e:
        tkinter.messagebox.showerror("Gerador de Planilha", f"Ocorreu uma exceção não tratada\n{e.__class__.__name__} - {e}")