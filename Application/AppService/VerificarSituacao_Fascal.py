import pandas as pd
import time
from tkinter import filedialog
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Filtro_Faturamento_Simples import *
from seleniumwire import webdriver
from selenium.webdriver.chrome.options import Options
import tkinter
from page_element import PageElement

class Login(PageElement):
    prestador_pj = (By.XPATH, '//*[@id="tipoAcesso"]/option[7]')
    usuario = (By.XPATH, '//*[@id="login-entry"]')
    senha = (By.XPATH, '//*[@id="password-entry"]')
    entrar = (By.XPATH, '//*[@id="BtnEntrar"]')

    def logar(self, usuario, senha):
        self.driver.find_element(*self.prestador_pj).click()
        self.driver.find_element(*self.usuario).send_keys(usuario)
        self.driver.find_element(*self.senha).send_keys(senha)
        self.driver.find_element(*self.entrar).click()
        time.sleep(10)

class Caminho(PageElement):
    localizar_procedimentos = (By.XPATH, '//*[@id="menuPrincipal"]/div/div[4]/a/span')
    Alerta = (By.XPATH, '/html/body/ul/li/div/div[2]/button[2]')

    def exe_caminho(self):
        try:
            element = WebDriverWait(self.driver, 20.0).until(EC.presence_of_element_located((By.XPATH, '//*[@id="menuPrincipal"]/div/div[4]/a/span')))
        except:
            print("Tempo de espera excedido. O site pode estar com delay ou fora do ar.")
        self.driver.find_element(*self.localizar_procedimentos).click()
        time.sleep(5)
        self.driver.find_element(*self.Alerta).click()
        time.sleep(5)
    
    
class injetar_dados(PageElement):
    guia_op = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/input') 
    buscar = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/span/span')
    
    def inserir_dados(self):
        faturas_df = pd.read_excel(planilha)
        count = 0
        for index, linha in faturas_df.iterrows():
            guia = (f"{linha['Nº Guia']}").replace(".0","")
            if (f"{linha['Pesquisado no Portal']}") == "Sim":
                print('Já foi feita a pesquisa desta autorização.')
                count = count + 1
                print(count)
                print('___________________________________________________________________________')
                continue
            self.driver.find_element(*self.guia_op).send_keys(guia)
            time.sleep(5)
            element = WebDriverWait(self.driver, 20.0).until(EC.presence_of_element_located((By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/span/span')))
            self.driver.find_element(*self.buscar).click()
            print("Pesquisa realizada")
            count = count + 1
            print(count)
            time.sleep(3)

            try:
                situacao = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div[3]/span').text
                dado = [situacao]
                pesquisado = 'Sim'
                print(f"{guia} está {dado}")
                dado_pesqu = {'Situação': [dado], 'Pesquisado no Portal': [pesquisado]}
                df = pd.DataFrame(dado_pesqu)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, "Sheet1", startrow= count, startcol=7, header=False, index=False)
                writer.save()
                self.driver.find_element(*self.guia_op).clear()
                time.sleep(2)
                print('___________________________________________________________________________')
            except:
                situacao = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[1]').text
                dado = [situacao]
                pesquisado = 'Sim'
                print(f"{guia}: {dado}")
                dado_pesqu = {'Situação': [dado], 'Pesquisado no Portal': [pesquisado]}
                df = pd.DataFrame(dado_pesqu)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, "Sheet1", startrow= count, startcol=7, header=False, index=False)
                writer.save()
                self.driver.find_element(*self.guia_op).clear()
                time.sleep(2)
                ('___________________________________________________________________________')
        self.driver.quit()


#-------------------------------------------------------------------------
def verificacao_fascal(user, password):
    try:

        try:
            processar_planilha()
            remove()
        except:
            pass

        planilha = filedialog.askopenfilename()

        url = 'https://novowebplanfascal.facilinformatica.com.br/GuiasTISS/Logon'

        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')

        options = {
        'proxy': {
                'http': f'http://{user}:{password}@10.0.0.230:3128',
                'https': f'http://{user}:{password}@10.0.0.230:3128'
            }
        }
        try:
            servico = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=servico, seleniumwire_options=options, options=chrome_options)
        except:
            driver = webdriver.Chrome(seleniumwire_options=options, options=chrome_options)
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro inesperado' )

    try:
        login_page = Login(driver, url)
        login_page.open()

        login_page.logar(
            usuario = '00735860000173',
            senha = '00735860000173'
            )

        time.sleep(4)

        Caminho(driver,url).exe_caminho()

        injetar_dados(driver,url).inserir_dados()

        print("Todas as guias foram pesquisadas com sucesso.")
        tkinter.messagebox.showinfo( 'Automação Faturamento - Fascal' , 'Buscas no portal da Fascal concluídos 😎✌' )
    except:
        tkinter.messagebox.showerror( 'Erro na busca' , 'Ocorreu um erro enquanto o Robô trabalhava, provavelmente o portal do Fascal caiu 😢' )
        driver.quit()