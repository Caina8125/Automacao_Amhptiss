import pandas as pd
import time
from tkinter import filedialog
from selenium import webdriver
from seleniumwire import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Filtro_Faturamento import processar_planilha, remove
from selenium.webdriver.chrome.options import Options
import tkinter
from page_element import PageElement

class Login(PageElement):
    prestador_pj = (By.XPATH, '//*[@id="tipoAcesso"]/option[9]')
    usuario = (By.XPATH, '//*[@id="login-entry"]')
    senha = (By.XPATH, '//*[@id="password-entry"]')
    entrar = (By.XPATH, '//*[@id="BtnEntrar"]')

    def logar(self, usuario, senha):
        time.sleep(2)
        self.driver.find_element(*self.prestador_pj).click()
        time.sleep(2)
        self.driver.find_element(*self.usuario).send_keys(usuario)
        time.sleep(2)
        self.driver.find_element(*self.senha).send_keys(senha)
        time.sleep(2)
        self.driver.find_element(*self.entrar).click()
        time.sleep(5)

class Caminho(PageElement):
    localizar_procedimentos = (By.XPATH, '//*[@id="menuPrincipal"]/div/div[2]/a/span')
    Alerta = (By.XPATH, '/html/body/ul/li/div/div[2]/button[2]')

    def exe_caminho(self):
        try:
            element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="menuPrincipal"]/div/div[4]/a/span')))
            self.driver.find_element(*self.localizar_procedimentos).click()
            time.sleep(2)
            self.driver.find_element(*self.Alerta).click()
            time.sleep(2)
        except:
            self.driver.refresh()
            login_page.logar(usuario = '00735860000173_2', senha = '00735860000173')
            self.driver.implicitly_wait(30)
            self.driver.find_element(*self.localizar_procedimentos).click()
            time.sleep(2)
            self.driver.find_element(*self.Alerta).click()
            time.sleep(2)
    
class injetar_dados(PageElement):
    guia_op = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/input') 
    buscar = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/span/span')

    def confere_planilha(self, novo_df, numero_guia, df):
        if novo_df.empty == True:
            novo_df = df.loc[(df["Nº Guia"] == int(numero_guia))]
            return novo_df
        else:
            return novo_df
    
    def inserir_dados(self):
        faturas_df = pd.read_excel(planilha)
        count = 0
        guia_loc = None
        
        for index, linha in faturas_df.iterrows():
            guia = str(linha['Nº Guia']).replace('.0', '')

            if not guia.isdigit():
                count += 1
                print(f'Nr. Guia {guia} é inválido')
                data = {'Situação': ['Número da guia operadora inválida(Possui letra)'], 'Validação Carteira': [''], 'Validação Proc.': [''], 'Validação Senha': [''], 'Pesquisado no Portal': ['Sim']}
                df = pd.DataFrame(data)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df.to_excel(writer, "Sheet1", startrow=count, startcol=7, header=False, index=False)
                writer.save()
                writer.close()
                continue   

            if linha['Pesquisado no Portal'] == "Sim":
                print('Já foi feita a pesquisa desta autorização.')
                count = count + 1
                print(count)
                print('___________________________________________________________________________')
                continue

            if guia_loc == guia:
                count = count + 1
                continue
            
            guia_loc = int(linha['Nº Guia'])
            pesquisa = False

            while pesquisa == False:

                try:
                    self.driver.find_element(*self.guia_op).clear() 
                    self.driver.find_element(*self.guia_op).send_keys(guia)
                    self.driver.find_element(*self.buscar).click()
                    pesquisa = True

                except:
                    pass

            user = False

            while user == False:

                try:
                    usuario = self.driver.find_element(By.XPATH, '//*[@id="menu_78B1E34CFC8E414D8EB4F83B534E4FB4"]').click()
                    user = True

                except:
                    pass
            
            self.driver.execute_script('scrollBy(0,1000)')
            self.driver.execute_script('scrollBy(0,1000)')
            time.sleep(1)

            count = count + 1

            guia_df = faturas_df.loc[(faturas_df["Nº Guia"] == guia)]
            guia_df = self.confere_planilha(guia_df, guia, faturas_df)
            count2 = 0

            for index2, linha2 in guia_df.iterrows():

                if linha['Pesquisado no Portal'] == "Sim":
                    print('Já foi feita a pesquisa desta autorização.')
                    count2 = count2 + 1
                    print(count)
                    print('___________________________________________________________________________')
                    continue
                
                try:
                    user = False

                    while user == False:

                        try:
                            usuario = self.driver.find_element(By.XPATH, '//*[@id="menu_78B1E34CFC8E414D8EB4F83B534E4FB4"]').click()
                            time.sleep(2)
                            user = True

                        except:
                            pass
                    self.driver.implicitly_wait(3)
                    situacao = WebDriverWait(self.driver, 3).until(EC.presence_of_element_located((By.XPATH,'//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div[3]/span'))).text
                    time.sleep(0.5)
                    print(f"{guia} está {situacao}")

                    carteira = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[1]/div[1]/strong[2]').text.replace("-", "")
                    carteira_planilha = (linha2['Matríc. Convênio']).replace('Nº - ', '')

                    procedimentos_planilha = f"{linha2['Procedimento']}".replace('.', '').replace('-', '')

                    if procedimentos_planilha == "10101012":
                        validacao_senha = "Senha não obrigatória"

                    else:
                        senha_portal = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[2]/div[1]/div[3]/span').text
                        time.sleep(0.5)
                        senha_planilha = f'{linha2["Senha Aut."]}'.replace(".0", "")

                        if senha_portal == senha_planilha:
                            validacao_senha = "Ok"

                        else:
                            validacao_senha = "Inválida"

                    if carteira == carteira_planilha:
                        matricula = 'Válida'
                    
                    else:
                        matricula = f'Inválida. Correta: {carteira}'

                    try:
                        self.driver.implicitly_wait(0.5)
                        procedimentos = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div[5]/a')
                        time.sleep(2)
                        procedimentos = procedimentos.get_attribute('outerHTML')
                        time.sleep(2)
                        procedimentos = procedimentos.replace('<a href="#" data-toggle="tooltip" data-placement="top" data-bind="attr: { title: $parent.CodigoAMB }" title="" data-original-title="', '')
                        time.sleep(2)
                        procedimentos = procedimentos.replace('-', '').replace('.', '')

                    except:
                        procedimentos = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[1]/div/div[2]/div/div').text.replace('-', '').replace('.', '')
                    
                    if procedimentos_planilha[0] == '1' and len(procedimentos_planilha) == 9:
                        data = {'Situação': [situacao], 'Validação Carteira': [matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Validação Senha': [validacao_senha], 'Pesquisado no Portal': ['Sim']}
                        df = pd.DataFrame(data)
                        book = load_workbook(planilha)
                        time.sleep(0.5)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                        writer.save()
                        writer.close()
                        count2 = count2 + 1
                        continue   

                    if procedimentos_planilha[0] == "0" or procedimentos_planilha[0] == "6" or procedimentos_planilha[0] == "7" or procedimentos_planilha[0] == "8":
                        data = {'Situação': [situacao], 'Validação Carteira': [matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Validação Senha': [validacao_senha], 'Pesquisado no Portal': ['Sim']}
                        df = pd.DataFrame(data)
                        book = load_workbook(planilha)
                        time.sleep(0.5)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                        writer.save()
                        writer.close()
                        count2 = count2 + 1
                        continue 

                    while procedimentos_planilha[0] == "9":
                        if procedimentos_planilha[1] == "8":
                            print("Procedimento")
                            break

                        else:
                            data = {'Situação': [situacao], 'Validação Carteira': [matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Validação Senha': [validacao_senha], 'Pesquisado no Portal': ['Sim']}
                            df = pd.DataFrame(data)
                            book = load_workbook(planilha)
                            time.sleep(0.5)
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                            df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                            writer.save()
                            writer.close()
                            count2 = count2 + 1
                            break

                    if procedimentos_planilha[0] == "9" and procedimentos_planilha[1] != "8":
                        continue

                    if procedimentos_planilha in procedimentos:
                        dados_proc = 'Ok'

                    else:
                        dados_proc = 'Não consta nesta autorização'
                    
                    data = {'Situação': [situacao], 'Validação Carteira': [matricula], 'Validação Proc.': [dados_proc], 'Validação Senha': [validacao_senha], 'Pesquisado no Portal': ['Sim']}
                    df = pd.DataFrame(data)
                    book = load_workbook(planilha)
                    time.sleep(0.5)
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                    writer.save()
                    writer.close()
                    count2 = count2 + 1
                    print('___________________________________________________________________________')
                except:
                    situacao = self.driver.find_element(By.XPATH, '//*[@id="localizarprocedimentos"]/div[2]/div/div[1]').text
                    data = {'Situação': [situacao], 'Validação Carteira': [''], 'Validação Proc.': [''], 'Validação Senha': [''], 'Pesquisado no Portal': ['Sim']}
                    print(f"{guia}: {situacao}")
                    df = pd.DataFrame(data)
                    book = load_workbook(planilha)
                    time.sleep(0.5)
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                    writer.save()
                    writer.close()
                    self.driver.find_element(*self.guia_op).clear()
                    ('___________________________________________________________________________')
            



#-------------------------------------------------------------------------

def verificacao_brb(user, password):
    try:
        try:
            processar_planilha()
            remove()
        except:
            pass
        global planilha
        planilha = filedialog.askopenfilename()
        url = 'https://portal.saudebrb.com.br/GuiasTISS/Logon'

        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')

        options = {
        'proxy': {
                'http': f'http://{user}:{password}@10.0.0.230:3128',
                'https': f'http://{user}:{password}@10.0.0.230:3128'
            }
        }
        try:
            servico = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=servico, options=chrome_options, seleniumwire_options=options)
        except:
            driver = webdriver.Chrome(options=chrome_options, seleniumwire_options=options)

        global login_page
        login_page = Login(driver, url)
        login_page.open()
        login_page.logar(
            usuario = '00735860000173_2',
            senha = 'AMHPDF00735@'
            )

        time.sleep(4)

        Caminho(driver,url).exe_caminho()

        injetar_dados(driver,url).inserir_dados()

        tkinter.messagebox.showinfo( 'Automação Faturamento - BRB' , 'Buscas no portal da BRB concluídos 😎✌' )

    except Exception as err:
        tkinter.messagebox.showerror( 'Erro na busca' , f'Ocorreu uma exceção não tratada \n {err.__class__.__name__} - {err}' )
        driver.quit()