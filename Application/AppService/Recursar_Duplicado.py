import pandas as pd
import time
from abc import ABC
from tkinter import filedialog
from selenium import webdriver
from seleniumwire import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import tkinter
from page_element import PageElement

class Login(PageElement):
    multiusuario = (By.XPATH, '/html/body/div[3]/div[3]/div/form/div[1]/label')
    prestador = (By.XPATH, '//*[@id="login_code"]')
    cpf = (By.XPATH, '//*[@id="login_cpf"]')
    senha = (By.XPATH, '//*[@id="login_password"]')
    logar = (By.XPATH, '//*[@id="btnLogin"]')

    def exe_login(self, prestador, cpf, senha):
        self.driver.find_element(*self.multiusuario).click()
        self.driver.find_element(*self.prestador).send_keys(prestador)
        self.driver.find_element(*self.cpf).send_keys(cpf)
        self.driver.find_element(*self.senha).send_keys(senha)
        self.driver.find_element(*self.logar).click()
        time.sleep(4)

class caminho(PageElement):
    Alerta = (By.XPATH, '/html/body/div[2]/div/center/a')
    botão_portaltiss = (By.XPATH, '//*[@id="main"]/div/div/div[2]/div[1]/nav/ul/li[21]/a')
    abrir_digitatiss = (By.XPATH, '/html/body/main/div/div/div[2]/div[2]/article/div[6]/div[4]/div[2]/div/div[2]/ul/li[1]/a')
    botão_recursoglosa = (By.XPATH, '//*[@id="HyperLink1"]')
    botão_novorecurso = (By.XPATH, '//*[@id="MenuOptionNew"]')

    def exe_caminho(self):
        # try:
        #     self.driver.find_element(*self.Alerta).click()
        # except:
        #     print('Não tem alerta')

        self.driver.find_element(*self.botão_portaltiss).click()
        time.sleep(2)
        self.driver.find_element(*self.abrir_digitatiss).click()
        time.sleep(2)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(1)
        self.driver.find_element(*self.botão_recursoglosa).click()
        time.sleep(1)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(1)
        self.driver.find_element(*self.botão_novorecurso).click()
        time.sleep(2)
        
class injetar_dados(PageElement):
    nro_guia = (By.XPATH, '//*[@id="NroGspPrincipal"]')
    baixar_guia = (By.XPATH, '//*[@id="btnQuickGuia"]')
    entrar_guia = (By.XPATH, '//*[@id="MenuOptionInsert"]')
    inserir_data = (By.XPATH, '//*[@id="DtaAtendimento"]')
    inserir_procedimento = (By.XPATH, '// *[ @ id = "NroServico"]')
    inserir_valor = (By.XPATH, '// *[ @ id = "VlrInformado"]')
    inserir_justificativa = (By.XPATH, '// *[ @ id = "ObsPrestador"]')
    salvar_recurso = (By.XPATH, '//*[@id="btn_gridProc18"]')
    voltar = (By.XPATH, '//*[@id="tb_gridProc18"]/tbody/tr[5]/td/span/a')
    grau_12 = (By.XPATH, '//*[@id="NroGrauParticipacao"]/option[6]')
    grau_00 = (By.XPATH, '//*[@id="NroGrauParticipacao"]/option[5]')
    grau_13 = (By.XPATH, '//*[@id="NroGrauParticipacao"]/option[9]')
    grau_1 = (By.XPATH, '//*[@id="NroGrauParticipacao"]/option[12]')
    grau_2 = (By.XPATH, '//*[@id="NroGrauParticipacao"]/option[14]')
    grau_3 = (By.XPATH, '//*[@id="NroGrauParticipacao"]/option[15]')
    abrir_envio = (By.XPATH,'//*[@id="tbInternalRow_NroDocumento"]/td[1]/span[2]/span[2]/a')
    opcao_enviar = (By.XPATH, '//*[@id="NroSitGspDigitacao"]/option[2]')
    enviar = (By.XPATH, '//*[@id="MenuOptionUpdate"]')
    sair = (By.XPATH, '//*[@id="MenuOptionExit"]')
    nova_guia = (By.XPATH, '//*[@id="MenuOptionNew"]')
    abrir_remessa = (By.XPATH, '//*[@id="lnkDT011"]')
    remessa_novo = (By.XPATH, '//*[@id="MenuOptionNew"]')
    # Esse botão é referente ao tipo "18-Recurso Glosa"
    tipo_remessa = (By.XPATH, '//*[@id="NroTpoGsp"]/option[8]')
    # Esse botão é referente a opção "Não"
    arquivos_anexados = (By.XPATH, '//*[@id="StaArquivoAnexado"]/option[2]')
    # Esse botão é referente a opção "Não"
    liberar_remessa = (By.XPATH, '//*[@id="NroConfRemessa"]/option[2]')
    # Esse botão é referente Gerar a Remessa
    botao_incluir = (By.XPATH, '//*[@id="MenuOptionInsert"]')

    def formatar(valor):
        return "{:,.2f}".format(valor)


    def inserir_dados(self):
        global count_linha
        count_linha = 0
        count_guia = 0
        global valor_total
        valor_total = 0

        faturas_df = pd.read_excel(planilha)
        try:
            faturas_df["Valor Recursado"] = faturas_df["Valor Recursado"].apply(injetar_dados.formatar).astype(str)
        except:
            pass
        for index, linha in faturas_df.iterrows():
            self.driver.find_element(*self.nro_guia).send_keys(f"{linha['Nro. Guia']}")
            time.sleep(1)
            self.driver.find_element(*self.baixar_guia).click()
            time.sleep(1)
            count_guia = count_guia + 1
            print('Guia', count_guia)
            break

        for index, linha in faturas_df.iterrows():
            valor_guia = float(f"{linha['Valor Soma']}")
            valor_total += valor_guia
            self.driver.find_element(*self.nro_guia).clear()
            time.sleep(1)
            self.driver.find_element(*self.nro_guia).send_keys(f"{linha['Nro. Guia']}")
            time.sleep(1)
            print(f"{linha['Nro. Guia']}", f"{linha['Procedimento']}", f"{linha['Realizado']}")
            self.driver.find_element(*self.baixar_guia).click()
            time.sleep(1)
            
            try:
                verifica_data = self.driver.find_element(By.XPATH, '//*[@id="tb_gridProc18"]/tbody/tr[1]/td[1]/table/tbody/tr/td[1]/strong').text
                if verifica_data == "Data Atendimento:":
                    print("Dentro da guia")
                    count_linha = count_linha + 1
                    print(count_linha)           
            except:
                self.driver.find_element(*self.entrar_guia).click()
                time.sleep(2)
                guia_geap = self.driver.find_element(By.XPATH, '//*[@id="NroGsp_fixed"]').text
                guia_planilha = [guia_geap]
                df = pd.DataFrame(guia_planilha)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                count_linha = count_linha + 1
                print(count_linha)
                df.to_excel(writer, "Recurso", startrow= count_linha, startcol=10, header=False, index=False)
                writer.save()
                time.sleep(1)
                print('anexado na planilha')

           
            time.sleep(3)
            self.driver.find_element(*self.inserir_data).send_keys(f"{linha['Realizado']}")
            self.driver.find_element(*self.inserir_procedimento).send_keys(f"{linha['Procedimento']}")
            self.driver.find_element(*self.inserir_valor).send_keys(f"{linha['Valor Recursado']}")
            self.driver.find_element(*self.inserir_justificativa).send_keys(f"{linha['Recurso Glosa']}")
            time.sleep(2)
        

            if (f"{linha['Grau Participação']}") == "12":
                self.driver.find_element(*self.grau_12).click()
                time.sleep(1)
                self.driver.find_element(*self.salvar_recurso).click()
                time.sleep(1)

            if (f"{linha['Grau Participação']}") == "0":
                self.driver.find_element(*self.grau_00).click()
                time.sleep(1)
                self.driver.find_element(*self.salvar_recurso).click()
                time.sleep(1)

            if (f"{linha['Grau Participação']}") == "13":
                self.driver.find_element(*self.grau_13).click()
                time.sleep(1)
                self.driver.find_element(*self.salvar_recurso).click()
                time.sleep(1)

            if (f"{linha['Grau Participação']}") == "1":
                self.driver.find_element(*self.grau_1).click()
                time.sleep(1)
                self.driver.find_element(*self.salvar_recurso).click()
                time.sleep(1)

            if (f"{linha['Grau Participação']}") == "2":
                self.driver.find_element(*self.grau_2).click()
                time.sleep(1)
                self.driver.find_element(*self.salvar_recurso).click()
                time.sleep(1)

            if (f"{linha['Grau Participação']}") == "3":
                self.driver.find_element(*self.grau_3).click()
                time.sleep(1)
                self.driver.find_element(*self.salvar_recurso).click()
                time.sleep(1)

            
            try:
                element = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="tb_gridProc18"]/tbody/tr[5]/td/span/a')))
                time.sleep(1)
                self.driver.find_element(*self.voltar).click()
                time.sleep(1)
                self.driver.find_element(*self.abrir_envio).click()
                time.sleep(2)
                self.driver.switch_to.window(self.driver.window_handles[-1])
                time.sleep(1)
                self.driver.find_element(*self.opcao_enviar).click()
                time.sleep(1)
                self.driver.find_element(*self.enviar).click()
                time.sleep(1)
                self.driver.find_element(*self.sair).click()
                time.sleep(1)
                self.driver.switch_to.window(self.driver.window_handles[-1])
                time.sleep(2)
                self.driver.find_element(*self.nova_guia).click()
                time.sleep(2)
                count_guia = count_guia + 1
                print('Guia', count_guia)
                if count_guia == 31:
                    valor_total = valor_total - valor_guia
                    injetar_dados(self.driver,url).Gerar_Remessa()
                    count_guia = 1
                    valor_total = valor_guia
                    

                self.driver.find_element(*self.nro_guia).send_keys(f"{linha['Nro. Guia']}")
                time.sleep(1)
                print(f"{linha['Nro. Guia']}", f"{linha['Procedimento']}", f"{linha['Realizado']}")
                self.driver.find_element(*self.baixar_guia).click()
                time.sleep(2)
                self.driver.find_element(*self.entrar_guia).click()
                time.sleep(2)
        

                guia_geap = self.driver.find_element(By.XPATH, '//*[@id="NroGsp_fixed"]').text
                guia_planilha = [guia_geap]
                df = pd.DataFrame(guia_planilha)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                #count_linha = count_linha + 1
                print(count_linha)
                df.to_excel(writer, "Recurso", startrow= count_linha, startcol=10, header=False, index=False)
                writer.save()
                time.sleep(1)
                print('anexado na planilha')
                try:
                    element = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="msgCheck"]')))
                    invalido = self.driver.find_element(By.XPATH, '//*[@id="msgCheck"]').text
                    if invalido == "Guia não localizada!!!":
                        self.driver.find_element(*self.nro_guia).clear()
                        time.sleep(1)
                        self.driver.find_element(*self.nro_guia).send_keys(f"{linha['Controle']}")
                        time.sleep(1)
                        self.driver.find_element(*self.baixar_guia).click()
                        time.sleep(1)
                        self.driver.find_element(*self.botao_incluir).click()
                        time.sleep(1)
                except:
                    print('Guia Valida')

                time.sleep(3)
                self.driver.find_element(*self.inserir_data).send_keys(f"{linha['Realizado']}")
                self.driver.find_element(*self.inserir_procedimento).send_keys(f"{linha['Procedimento']}")
                self.driver.find_element(*self.inserir_valor).send_keys(f"{linha['Valor Recursado']}")
                self.driver.find_element(*self.inserir_justificativa).send_keys(f"{linha['Recurso Glosa']}")

                if (f"{linha['Grau Participação']}") == "12":
                    self.driver.find_element(*self.grau_12).click()
                    time.sleep(1)
                    self.driver.find_element(*self.salvar_recurso).click()
                    time.sleep(1)

                if (f"{linha['Grau Participação']}") == "0":
                    self.driver.find_element(*self.grau_00).click()
                    time.sleep(1)
                    self.driver.find_element(*self.salvar_recurso).click()
                    time.sleep(1)

                if (f"{linha['Grau Participação']}") == "13":
                    self.driver.find_element(*self.grau_13).click()
                    time.sleep(1)   
                    self.driver.find_element(*self.salvar_recurso).click()  
                    time.sleep(1)

                if (f"{linha['Grau Participação']}") == "1":
                    self.driver.find_element(*self.grau_1).click()
                    time.sleep(1)
                    self.driver.find_element(*self.salvar_recurso).click()
                    time.sleep(1)

                if (f"{linha['Grau Participação']}") == "2":
                    self.driver.find_element(*self.grau_2).click()
                    time.sleep(1)
                    self.driver.find_element(*self.salvar_recurso).click()
                    time.sleep(1)

                if (f"{linha['Grau Participação']}") == "3":
                    self.driver.find_element(*self.grau_3).click()
                    time.sleep(1)
                    self.driver.find_element(*self.salvar_recurso).click()
                    time.sleep(1)
                continue

            except:
                guia_geap = self.driver.find_element(By.XPATH, '//*[@id="NroGsp_fixed"]').text
                guia_planilha = [guia_geap]
                df = pd.DataFrame(guia_planilha)
                book = load_workbook(planilha)
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                #count_linha = count_linha + 1
                #print(count_linha)
                df.to_excel(writer, "Recurso", startrow= count_linha, startcol=10, header=False, index=False)
                writer.save()
                time.sleep(1)
                print('anexado na planilha')
                continue

        self.driver.find_element(*self.abrir_envio).click()
        time.sleep(2)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(1)
        self.driver.find_element(*self.opcao_enviar).click()
        time.sleep(1)
        self.driver.find_element(*self.enviar).click()
        time.sleep(1)
        self.driver.find_element(*self.sair).click()
        time.sleep(1)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(2)
        self.driver.find_element(*self.nova_guia).click()
        count_linha = count_linha + 1        
        injetar_dados(self.driver,url).Gerar_Remessa()
        self.driver.quit()



    def Gerar_Remessa(self):
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(1)
        self.driver.close()
        time.sleep(1)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(1)
        self.driver.find_element(*self.abrir_remessa).click()
        time.sleep(1)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(1)
        self.driver.find_element(*self.remessa_novo).click()
        time.sleep(1)

        # Selecionar o tipo de Remessa "18 - Recurso de Glosa"
        self.driver.find_element(*self.tipo_remessa).click()
        time.sleep(1)
        # Na opção de arquivos anexados selecionar "Não"
        self.driver.find_element(*self.arquivos_anexados).click()
        time.sleep(1)
        # Na opção de liberar remessa após a geração, marcar "Não"
        self.driver.find_element(*self.liberar_remessa).click()
        time.sleep(2)
        self.driver.find_element(*self.botao_incluir).click()
        time.sleep(3)
        self.driver.switch_to.window(self.driver.window_handles[-1])
        time.sleep(1)
        n_remessa = self.driver.find_element(By.XPATH, '//*[@id="objTableHeader"]/tbody/tr[1]/td[2]').text
        self.driver.save_screenshot(f"{n_remessa}.png")
        remessa_valor = {'Remessa': [n_remessa], 'Valor Total': [valor_total]}
        df = pd.DataFrame(remessa_valor)
        book = load_workbook(planilha)
        writer = pd.ExcelWriter(planilha, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, "Recurso", startrow = count_linha - 1, startcol=11, header=False, index=False)
        writer.save()
        time.sleep(2)
        print(planilha)
        time.sleep(2)
        self.driver.get('https://www2.geap.com.br/digitaTiss/DT001_GUIA_18.aspx')
        time.sleep(1)
        self.driver.find_element(By.XPATH, '//*[@id="MenuOptionNew"]').click()


#---------------------------------------------------------------------------------
def recursar_duplicado(user, password):
    global planilha
    planilha = filedialog.askopenfilename()

    global url
    url = 'https://www2.geap.com.br/auth/prestador.asp'

    servico = Service(ChromeDriverManager().install())

    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--ignore-ssl-errors')

    options = {
    'proxy': {
            'http': f'http://{user}:{password}@10.0.0.230:3128',
                'https': f'http://{user}:{password}@10.0.0.230:3128'
        }
    }
    try:
        driver = webdriver.Chrome(service=servico, seleniumwire_options=options, options=chrome_options)
    except:
        driver = webdriver.Chrome(seleniumwire_options=options, options=chrome_options)

    try:
        login_page = Login(driver, url)
        login_page.open()
        login_page.exe_login(
            prestador = "23003723",
            cpf = '66661692120',
            senha = "amhpdf0073"
        )

        time.sleep(4)

        caminho(driver,url).exe_caminho()

        injetar_dados(driver,url).inserir_dados()
        tkinter.messagebox.showinfo( 'Automação GEAP Recurso de Glosa' , 'Recursos na GEAP Concluídos 😎✌' )
    
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro enquanto o Robô trabalhava, provavelmente o portal da GEAP caiu 😢' )
        driver.quit()