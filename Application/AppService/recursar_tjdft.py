from tkinter import filedialog
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium import webdriver
from openpyxl import load_workbook
import pandas as pd
import time
import os
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from seleniumwire import webdriver
import tkinter
from page_element import PageElement


class Login(PageElement):
    email = (By.XPATH, '//*[@id="Email"]')
    senha = (By.XPATH, '//*[@id="Senha"]')
    logar = (By.XPATH, '//*[@id="btnLogin"]')

    def exe_login(self, email, senha):
        self.driver.find_element(*self.email).send_keys(email)
        time.sleep(2)
        self.driver.find_element(*self.senha).send_keys(senha)
        time.sleep(2)
        self.driver.find_element(*self.logar).click()


class caminho(PageElement):
    lote_de_pagamento = (By.XPATH, '/html/body/div[3]/div[1]/div/ul/li[17]/a/span[1]')
    pesquisar_lotes = (By.XPATH, '/html/body/div[3]/div[1]/div/ul/li[17]/ul/li[3]/a/span[1]')
    proximo = (By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[2]')
    fechar = (By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[3]')
    fechar_botao = (By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[2]')
    fechar_alerta = (By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')

    def exe_caminho(self):
        try:
            modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div')))
            self.driver.find_element(*self.fechar_botao).click()

            while True:
                try:
                    proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[2]')))
                    proximo_botao.click()
                except:
                    break

            try:
                fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                fechar_botao.click()
            except:
                print("Não foi possível encontrar o botão de fechar.")
                pass

        except:
            print("Não teve Modal")
            pass
        self.driver.find_element(*self.lote_de_pagamento).click()
        time.sleep(2)
        self.driver.find_element(*self.pesquisar_lotes).click()
        time.sleep(3)
        try:
            modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div')))
            self.driver.find_element(*self.fechar_botao).click()

            while True:
                try:
                    proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[2]')))
                    proximo_botao.click()
                except:
                    break

            try:
                fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                fechar_botao.click()
            except:
                print("Não foi possível encontrar o botão de fechar.")
                pass

        except:
            print("Não teve Modal")
            pass

    def novo_rec(self):
        self.driver.find_element(*self.pesquisar_lotes).click()
        print('Dentro de novo recurso')

    def Alert(self):
        self.driver.find_element(*self.proximo).click()
        time.sleep(1)
        self.driver.find_element(*self.fechar).click()                    

class inserir_dados(PageElement):
    protocolo = (By.XPATH, '//*[@id="Protocolo"]')
    pesquisar = (By.XPATH, '//*[@id="btn-Pesquisar"]/span')
    selecionar = (By.XPATH, '//*[@id="DataGrid"]/tbody[1]/tr[1]/td[2]/input')
    atualizar = (By.XPATH, '/html/body/div[3]/div[2]/div/div[2]/div/div[2]/div/div/div[2]/div/bc-smart-table/div[2]/table/thead/tr[1]/th/div[1]/div/div/a[3]/span')
    recursoclick = (By.XPATH, '//*[@id="recurso-glosa"]/span')
    controle = (By.XPATH, '//*[@id="GuiasGlosadasTable"]/thead/tr[1]/th/div[2]/input')
    alerta2 = (By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[3]')
    marcar = (By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]/input')
    procedimento = (By.XPATH, '//*[@id="bcRecursar"]/span')
    checkbox = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[1]/input')
    checkbox2 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[1]/input')
    checkbox3 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[1]/input')
    checkbox4 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[1]/input')
    checkbox5 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[1]/input')
    checkbox6 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[1]/input')
    checkbox7 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[1]/input')
    checkbox8 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[1]/input')
    checkbox9 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[1]/input')
    checkbox10 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[1]/input')
    recursar = (By.LINK_TEXT, 'Recursar Procedimento')
    justificativas = (By.CLASS_NAME, 'form-control ng-pristine ng-untouched ng-valid ng-empty ng-valid-maxlength')
    valor = (By.XPATH, '//*[@id="ValorRecursado"]')
    novo_recurso = (By.XPATH, '//*[@id="simpleConfirmationModal_btOk"]')
    pesquisar_proc = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/thead/tr[1]/th/div[2]/input')
    salvar = (By.XPATH, '/html/body/bc-modal-evolution/bc-modal-justificar-glosa/div/div/div/div[3]/button[2]')
    fechar_recurso = (By.XPATH, '/html/body/bc-modal-evolution/bc-modal-justificar-glosa/div/div/div/div[1]/button')
    fechar_botao = (By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[3]')
    fechar_alerta = (By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')

    def justificar(self, qtd_glosas, justificativa_plan):
        for i in range(0, qtd_glosas):
            justificativa = self.driver.find_element(By.ID, f'Justificativas[{i}].Justificativa')
            justificativa.click()
            time.sleep(1)
            justificativa.send_keys(justificativa_plan)
            time.sleep(2)

    def Protocolo(self):
        nomesarquivos = os.listdir(pasta)
        for nome in nomesarquivos:
            if "Enviado" in nome:
                print("PEG já enviado")
                continue
            sem_extensao = nome.replace('.xlsx', '')
            planilha = os.path.join(pasta, nome)
            faturas_df1 = pd.read_excel(planilha)
            for index, linha in faturas_df1.iterrows():
                try:
                    modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div')))
                    self.driver.find_element(*self.fechar_botao).click()

                    while True:
                        try:
                            proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[2]')))
                            proximo_botao.click()
                        except:
                            break

                    try:
                        fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                        fechar_botao.click()
                    except:
                        print("Não foi possível encontrar o botão de fechar.")
                        pass

                except:
                    print("Não teve Modal")
                    pass
                self.driver.find_element(*self.protocolo).send_keys(f"{linha['Protocolo Aceite']}")
                print('Protocolo: ' + f"{linha['Protocolo Aceite']}")
        
                time.sleep(1)
                self.driver.find_element(*self.pesquisar).click()
                time.sleep(2)
                try:
                    for contador in range(3):
                        self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
                except:
                    None

                try:
                    modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div')))
                    self.driver.find_element(*self.fechar_botao).click()

                    while True:
                        try:
                            proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[2]')))
                            proximo_botao.click()
                        except:
                            break

                    try:
                        fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                        fechar_botao.click()
                    except:
                        print("Não foi possível encontrar o botão de fechar.")
                        pass

                except:
                    print("Não teve Modal")
                    pass

                self.driver.find_element(*self.selecionar).click()
                
                time.sleep(2)
                situacao = self.driver.find_element(By.XPATH,'//*[@id="DataGrid"]/tbody[1]/tr[1]/td[7]').text
                if situacao == "Faturado":
                    self.driver.find_element(*self.recursoclick).click()
                    time.sleep(6)
                else:
                    self.driver.find_element(*self.atualizar).click()
                    time.sleep(6)
                    self.driver.find_element(*self.pesquisar).click()
                    time.sleep(2)
                    try:
                        for contador in range(3):
                            self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
                    except:
                        None
                    self.driver.find_element(*self.selecionar).click()
                    time.sleep(2)
                    self.driver.find_element(*self.recursoclick).click()
                    time.sleep(6)

                try:
                    element = WebDriverWait(self.driver, 50).until(EC.presence_of_element_located((self.novo_recurso)))
                    self.driver.find_element(*self.novo_recurso).click()
                except:
                    print('Recurso Normal')
                time.sleep(2)
                try:
                    element = WebDriverWait(self.driver, 50).until(EC.presence_of_element_located((self.controle)))
                    self.driver.find_element(*self.controle).clear()
                except:
                    print('Limpar Nº Guia')
                try:
                    self.driver.find_element(*self.alerta2).click()
                except:
                    print('Alerta2 não apareceu')
                break

            try:
                modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div')))
                self.driver.find_element(*self.fechar_botao).click()

                while True:
                    try:
                        proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[2]')))
                        proximo_botao.click()
                    except:
                        break

                try:
                    fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                    fechar_botao.click()
                except:
                    print("Não foi possível encontrar o botão de fechar.")
                    pass

            except:
                print("Não teve Modal")
                pass
            count = 0
            faturas_df1 = pd.read_excel(planilha)
                # break

            for index, linha in faturas_df1.iterrows():
                self.driver.find_element(*self.controle).clear()
                time.sleep(1)
                self.driver.find_element(*self.controle).send_keys(f"{linha['Controle Inicial']}")
                time.sleep(1)
                self.driver.find_element(*self.marcar).click()
                time.sleep(1)
                self.driver.find_element(*self.procedimento).click()
                time.sleep(2)
                self.driver.find_element(*self.pesquisar_proc).clear()
                time.sleep(2) 
                self.driver.find_element(*self.pesquisar_proc).send_keys(f"{linha['Procedimento']}")
                time.sleep(2)
                for t in range(10):
                    try:
                        self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/thead/tr[2]/th[8]').click()
                        break
                    except:
                        time.sleep(2)
                time.sleep(2)
                self.driver.find_element(*self.pesquisar_proc).clear()
                time.sleep(2)
                
                desmarcar = WebDriverWait(self.driver, 50).until(EC.presence_of_element_located((self.marcar)))
                time.sleep(1)
                self.driver.find_element(*self.controle).clear()
                time.sleep(1)
                self.driver.find_element(*self.controle).send_keys(f"{linha['Controle Inicial']}")
                self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]/input').click()
                time.sleep(2)
                self.driver.find_element(*self.controle).clear()
                print( 'Procedimentos ordenados')
                break

            for index, linha in faturas_df1.iterrows():
                print('---------------------------------------------------------------------------------------------------------')
                count = count + 1
                print(count)
                print('Controle: ' + f"{linha['Controle Inicial']}")

                if (f"{linha['Recursado no Portal']}") == "Sim":
                    print('Data: ' + f"{linha['Realizado']}")
                    print('Código: ' + f"{linha['Procedimento']}")
                    print('Procedimento já foi recursado no portal')
                    continue

                print('Data: ' + f"{linha['Realizado']}")
                time.sleep(2)
                self.driver.find_element(*self.controle).clear()
                self.driver.find_element(*self.controle).send_keys(f"{linha['Controle Inicial']}")
                
                time.sleep(2)
                self.driver.find_element(*self.marcar).click()

                time.sleep(1)
                self.driver.find_element(*self.procedimento).click()
                time.sleep(2)
                self.driver.find_element(*self.pesquisar_proc).clear()
                time.sleep(2) 
                self.driver.find_element(*self.pesquisar_proc).send_keys(f"{linha['Procedimento']}")
                print('Código: ' + f"{linha['Procedimento']}")
                
                time.sleep(3)
                
                try:
                    for contador in range(3):
                        self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
                except:
                    None
                try:
                    for i in range(0,10):
                        try:
                            valor1 =  self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[8]').text
                            codigo1 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[2]').text
                            glosas1 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[4]').text.split(', ')
                            for k, v in enumerate(codigo1):
                                if codigo1[0] != "0":
                                    break
                                if codigo1[k] != "0":
                                    codigo1 = str(codigo1[k:])
                                    break
                            vlr_glosado1 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[7]').text
                            elementos_encontrados = True
                            print("Elementos Encontrados")
                            break
                        except:
                            time.sleep(2)

                    valor_planilha = f"{linha['Valor Glosa']}"
                    procedimento_plan = f"{linha['Procedimento']}"

                    if ((valor1 == "R$0,00") and (codigo1 == procedimento_plan)) and (vlr_glosado1 == valor_planilha):
                        print('Valor do 1° Procedimento Encontrado = R$0,00')
                        print('Valor do 1° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)

                        self.driver.find_element(*self.checkbox).click()
                        time.sleep(1)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(1)
                        try:
                            button = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            time.sleep(1)
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas1), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas1), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(2)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(2)
                            print('Exceção')
                            continue

                    valor2 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[8]').text
                    codigo2 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[2]').text
                    glosas2 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo2):
                        if codigo2[0] != "0":
                            break
                        if codigo2[k] != "0":
                            codigo2 = str(codigo2[k:])
                            break
                    vlr_glosado2 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[7]').text

                    if ((valor2 == "R$0,00") and (codigo2 == f"{linha['Procedimento']}",".0") and (vlr_glosado2 >= "R$0,22")) and vlr_glosado2 == valor_planilha:
                        print('Valor do 2° Procedimento Encontrado = R$0,00')
                        print('Valor do 2° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox2).click()
                        time.sleep(1)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas2), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas2), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox2).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox2).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor3 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[8]').text
                    codigo3 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[2]').text
                    glosas3 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo3):
                        if codigo3[0] != "0":
                            break
                        if codigo3[k] != "0":
                            codigo3 = str(codigo3[k:])
                            break
                    vlr_glosado3 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[7]').text

                    if ((valor3 == "R$0,00") and (codigo3 == f"{linha['Procedimento']}",".0") and (vlr_glosado3 >= "R$0,22")) and vlr_glosado3 == valor_planilha:
                        print('Valor do 3° Procedimento Encontrado = R$0,00')
                        print('Valor do 3° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox3).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas3), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas3), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']
                            
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox3).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox3).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor4 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[8]').text
                    codigo4 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[2]').text
                    glosas4 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo4):
                        if codigo4[0] != "0":
                            break
                        if codigo4[k] != "0":
                            codigo4 = str(codigo4[k:])
                            break
                    vlr_glosado4 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[7]').text

                    if ((valor4 == "R$0,00") and (codigo4 == f"{linha['Procedimento']}",".0") and (vlr_glosado4 >= "R$0,22")) and vlr_glosado4 == valor_planilha:
                        print('Valor do 4° Procedimento Encontrado = R$0,00')
                        print('Valor do 4° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox4).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas4), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas4), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox4).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox4).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor5 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[8]').text
                    codigo5 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[2]').text
                    glosas5 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo5):
                        if codigo5[0] != "0":
                            break
                        if codigo5[k] != "0":
                            codigo5 = str(codigo5[k:])
                            break
                    vlr_glosado5 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[7]').text

                    if ((valor5 == "R$0,00") and (codigo5 == f"{linha['Procedimento']}",".0") and (vlr_glosado5 >= "R$0,22")) and vlr_glosado5 == valor_planilha:
                        print('Valor do 5° Procedimento Encontrado = R$0,00')
                        print('Valor do 5° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox5).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas5), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas5), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox5).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox5).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor6 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[8]').text
                    codigo6 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[2]').text
                    glosas6 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo6):
                        if codigo6[0] != "0":
                            break
                        if codigo6[k] != "0":
                            codigo6 = str(codigo6[k:])
                            break
                    vlr_glosado6 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[7]').text

                    if ((valor6 == "R$0,00") and (codigo6 == f"{linha['Procedimento']}",".0") and (vlr_glosado6 >= "R$0,22")) and vlr_glosado6 == valor_planilha:
                        print('Valor do 6° Procedimento Encontrado = R$0,00')
                        print('Valor do 6° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox6).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas6), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas6), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox6).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox6).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                        

                    valor7 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[8]').text
                    codigo7 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[2]').text
                    glosas7 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo7):
                        if codigo7[0] != "0":
                            break
                        if codigo7[k] != "0":
                            codigo7 = str(codigo7[k:])
                            break
                    vlr_glosado7 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[7]').text

                    if ((valor7 == "R$0,00") and (codigo7 == f"{linha['Procedimento']}",".0") and (vlr_glosado7 >= "R$0,22")) and vlr_glosado7 == valor_planilha:
                        print('Valor do 7° Procedimento Encontrado = R$0,00')
                        print('Valor do 7° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox7).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas7), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas7), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox7).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox7).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor8 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[8]').text
                    codigo8 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[2]').text
                    glosas8 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo8):
                        if codigo8[0] != "0":
                            break
                        if codigo8[k] != "0":
                            codigo8 = str(codigo8[k:])
                            break
                    vlr_glosado8 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[7]').text

                    if ((valor8 == "R$0,00") and (codigo8 == f"{linha['Procedimento']}",".0") and (vlr_glosado8 >= "R$0,22")) and vlr_glosado8 == valor_planilha:
                        print('Valor do 8° Procedimento Encontrado = R$0,00')
                        print('Valor do 8° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox8).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas8), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas8), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox8).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox8).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor9 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[8]').text
                    codigo9 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[2]').text
                    glosas9 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo9):
                        if codigo9[0] != "0":
                            break
                        if codigo9[k] != "0":
                            codigo9 = str(codigo9[k:])
                            break
                    vlr_glosado9 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[7]').text

                    if ((valor9 == "R$0,00") and (codigo9 == f"{linha['Procedimento']}",".0") and (vlr_glosado9 >= "R$0,22")) and vlr_glosado9 == valor_planilha:
                        print('Valor do 9° Procedimento Encontrado = R$0,00')
                        print('Valor do 9° Procedimento injetado = ' + f"{linha['Valor Recursado']}")
                        time.sleep(1)
                        self.driver.find_element(*self.checkbox9).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas9), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas9), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox9).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox9).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor10 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[8]').text
                    codigo10 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[2]').text
                    glosas10 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[4]').text.split(', ')
                    for k, v in enumerate(codigo10):
                        if codigo10[0] != "0":
                            break
                        if codigo10[k] != "0":
                            codigo10 = str(codigo10[k:])
                            break
                    vlr_glosado10 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[7]').text

                    if ((valor10 == "R$0,00") and (codigo10 == f"{linha['Procedimento']}",".0") and (vlr_glosado10 >= "R$0,22")) and vlr_glosado10 == valor_planilha:
                        print('Valor do 10° Procedimento Encontrado = R$0,00')
                        print('Valor do 10° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox10).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas10), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.justificar(len(glosas10), f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)

                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow = count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox10).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox10).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue
                except Exception as e:
                    print('Procedimento já recursados ou não existe esse código nesse protocolo.')
                    desmarcar = WebDriverWait(self.driver, 50).until(EC.presence_of_element_located((self.marcar))).click()
                    continue
            novo_nome = pasta + '/' + sem_extensao + '_Enviado.xlsx'
            try:
                writer.close()
                book.close()
            except:
                pass
            try:
                os.rename(planilha, novo_nome)
            except:
                print("Erro ao renomear arquivo")
            self.driver.get('https://prosaudeconecta.tjdft.jus.br/Pagamentos/PesquisaLote/Index')
        self.driver.quit()
#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def recursar_tjdft(user, password):
    try:
        url = 'https://prosaudeconecta.tjdft.jus.br/Account/Login'

        global pasta
        pasta = filedialog.askdirectory()

        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')

        options = {
        'proxy': {
                'http': f'http://{user}:{password}@10.0.0.230:3128',
                'https': f'http://{user}:{password}@10.0.0.230:3128'
            }
        }
        try:
            servico = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=servico, seleniumwire_options=options, options=chrome_options)
        except:
            driver = webdriver.Chrome(seleniumwire_options=options, options=chrome_options)
    except:
        tkinter.messagebox.showerror( 'Erro Automação' , 'Ocorreu um erro inesperado' )

    try:
        login_page = Login(driver, url)
        login_page.open()

        login_page.exe_login(
            email="recursoadministrativo@amhp.com.br", 
            senha="@Amhp19"
        )

        print('Pegar Alerta Acionado!')
        caminho(driver, url).exe_caminho()
        inserir = inserir_dados(driver, url)
        print('Protocolo Acionado')
        inserir.Protocolo()
        print('Todos os procedimentos foram recursados com sucesso.')
        tkinter.messagebox.showinfo( 'Automação' , 'Recursos do TJDFT Concluídos' )
    
    except Exception as e:
        tkinter.messagebox.showerror( 'Erro Automação' , f'Ocorreu uma excessão não tratada:\n{e.__class__.__name__} - {e}' )
        driver.quit()