import pandas as pd
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook
import os
from Application.AppService.AutomacaoService.page_element import PageElement

class RecursoStm(PageElement):
    body = (By.XPATH, '/html/body')
    medico = (By.XPATH, '//*[@id="tipoAcesso"]/option[9]')
    usuario_input = (By.XPATH, '//*[@id="login-entry"]')
    senha_input = (By.XPATH, '//*[@id="password-entry"]')
    entrar = (By.XPATH, '//*[@id="BtnEntrar"]')
    continuar = (By.XPATH, '/html/body/div[3]/div[3]/div/button[2]/span')
    fechar = (By.XPATH, '/html/body/ul/li/div/div/span/h4/i')
    faturas = (By.XPATH, '//*[@id="menuPrincipal"]/div/div[8]/a')
    relatorio_de_faturas = (By.XPATH, '/html/body/header/div[4]/div/div/div/div[8]/div[1]/div[2]/div/div[2]/div/div/div/div[1]/a')
    body = (By.XPATH, '/html/body')
    fechar = (By.XPATH, '/html/body/ul/li/div/div/span/h4/i')
    body = (By.XPATH, '/html/body')
    codigo = (By.XPATH, '/html/body/main/div/div[1]/div[2]/div/div/div[2]/div[1]/div[1]/input-text[1]/div/div/input')
    pesquisar = (By.XPATH, '//*[@id="filtro"]/div[2]/div[2]/button')
    recurso_de_glosa = (By.XPATH, '//*[@id="div-Servicos"]/div[1]/div[4]/div/div/div[1]/div/div[2]/a[5]/i')
    #-----------------------------------------------------------------------------------------------------------------------------
    table = (By.ID, 'recursoGlosaTabelaServicos')
    text_area_justificativa = (By.ID, 'txtJustificativa')
    #----------------------------------------------------------------------------------------------------------------------------------
    button_ok = ()
    salvar_parcialmente = ()
    i_close = ()
    proxima_pagina = ()
    label_registros = ()
    primeira_pagina = ()
    fechar = ()
    ul = ()
    #-----------------------------------------------------------------------------------------------------------------------------------------------
    close_warning = (By.XPATH, '/html/body/ul/li/div/div/span/h4/i')
    recurso_de_glosa_menu = (By.XPATH, '//*[@id="menuPrincipal"]/div/div[7]/a')
    fatura_input = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/input-number/div/div/input')
    pesquisar_recurso = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[2]/button[1]')
    recurso_de_glosa_2 = (By.XPATH, '/html/body/main/div[1]/div[2]/div/table/tbody/tr/td[11]/i')
    alerta = (By.XPATH, '/html/body/ul/li/div/div[2]/button[2]')
    guia_op = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/input') 
    buscar = (By.XPATH, '/html/body/main/div[1]/div[1]/div[2]/div[1]/div[2]/input-text-search/div/div/div/span/span')

    def __init__(self, url, usuario, senha):
        super().__init__(url)
        self.usuario = usuario
        self.senha = senha

    def exe_login(self):
        self.driver.implicitly_wait(30)
        self.driver.find_element(*self.medico).click()
        time.sleep(1.5)
        self.driver.find_element(*self.usuario_input).clear()
        self.driver.find_element(*self.usuario_input).send_keys(self.usuario)
        time.sleep(1.5)
        self.driver.find_element(*self.senha_input).send_keys(self.senha)
        time.sleep(1.5)
        self.driver.find_element(*self.entrar).click()
        time.sleep(1.5)
        self.driver.find_element(*self.continuar).click()
        time.sleep(5)
        content = self.driver.find_element(*self.body).text

        if "Sua sessão expirou." in content:
            self.driver.find_element(*self.fechar).click()
            time.sleep(1.5)
            self.driver.find_element(*self.medico).click()
            time.sleep(1.5)
            self.driver.find_element(*self.usuario_input).send_keys(self.usuario)
            time.sleep(1.5)
            self.driver.find_element(*self.senha_input).send_keys(self.senha)
            time.sleep(1.5)
            self.driver.find_element(*self.entrar).click()
            time.sleep(1.5)
            self.driver.find_element(*self.continuar).click()
            time.sleep(5)

    def exe_caminho(self):
        content = self.driver.find_element(*self.body).text
        if "Sua sessão expirou." in content:
            self.driver.find_element(*self.fechar).click()
            time.sleep(2)
            self.exe_login()

        if "Guias" not in content:
            self.driver.back()
            self.exe_login()
        try:
            self.driver.implicitly_wait(10)
            time.sleep(2)
            self.driver.find_element(*self.faturas)
        except:
            self.driver.refresh()
            time.sleep(2)
            self.exe_login()

        self.driver.implicitly_wait(30)
        time.sleep(3)
        self.driver.find_element(*self.faturas).click()
        time.sleep(2)
        self.driver.find_element(*self.relatorio_de_faturas).click()
        time.sleep(2)

    def inicializar_atributos(self, recurso_iniciado):
        if recurso_iniciado == False:
            self.button_ok = (By.XPATH, '/html/body/main/div/div[3]/div/div/div[3]/button[1]')
            self.salvar_parcialmente = (By.XPATH, '/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[2]/button')
            self.i_close = (By.XPATH, '/html/body/ul/li/div/div/span/h4/i')
            # self.proxima_pagina = (By.XPATH, '/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/nav/ul/li[6]/a/span')
            self.label_registros = (By.XPATH, '/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/label')
            self.primeira_pagina = (By.XPATH, '/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/nav/ul/li[1]/a/span')
            self.fechar = (By.XPATH, '/html/body/main/div/div[2]/div/div/div[3]/button')
            self.ul = (By.XPATH, '/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/nav/ul')
        else:
            self.button_ok = (By.XPATH, '/html/body/main/div[1]/div[4]/div/div/div[3]/button[1]')
            self.salvar_parcialmente = (By.XPATH, '/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[2]/button')
            self.i_close = (By.XPATH, '/html/body/ul/li/div/div/span/h4/i')
            # self.proxima_pagina = (By.XPATH, '/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/nav/ul/li[6]/a/span')
            self.label_registros = (By.XPATH, '/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/label')
            self.primeira_pagina = (By.XPATH, '/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/nav/ul/li[1]/a/span')
            self.fechar = (By.XPATH, '/html/body/main/div[1]/div[3]/div/div/div[3]/button')
            self.ul = (By.XPATH, '/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/nav/ul')

    def get_values(self, i, recurso_iniciado):
        if recurso_iniciado == False:
                for j in range(10):
                    try:
                        nro_guia_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[2]/span').text
                        codigo_proc_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[4]/span').text.replace('.', '').replace('-', '')
                        valor_glosa_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[6]/span').text
                        valor_recursado_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[8]/span').text
                        nome_paciente_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[3]/span').text.replace('-', '')
                        break
                    except:
                        print("Variáveis não encontradas.")
                        time.sleep(2)
                        continue

        else:
            for j in range(10):
                try:
                    nro_guia_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[2]/span').text
                    codigo_proc_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[4]/span').text.replace('.', '').replace('-', '')
                    valor_glosa_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[6]/span').text
                    valor_recursado_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[8]/span').text
                    nome_paciente_portal = self.driver.find_element(By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[3]/span').text.replace('-', '')
                    break
                except:
                    print("Variáveis não encontradas.")
                    time.sleep(2)
                    continue
        return (nro_guia_portal, codigo_proc_portal, valor_glosa_portal, valor_recursado_portal, nome_paciente_portal)
    
    def xpath_preencher_valores(self, i, recurso_iniciado):
        if recurso_iniciado == False:
            input_valor_recursado = (By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[7]/input')
            preencher_justificativa = (By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[9]/a/i')
        else:
            input_valor_recursado = (By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[7]/input')
            preencher_justificativa = (By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[9]/a/i')
        return (input_valor_recursado, preencher_justificativa)

    def inicia_automacao(self, **kwargs):
        self.init_driver()
        self.open()
        self.exe_login()
        self.exe_caminho()

        diretorio = kwargs.get('diretorio')

        count = 0
        while count < 10:
            try:
                lista_de_planilhas = [f"{diretorio}//{arquivo}" for arquivo in os.listdir(diretorio) if arquivo.endswith(".xlsx")]

                for planilha in lista_de_planilhas:
                    if "Enviado" in planilha or "Sem_Pagamento" in planilha:
                        continue
                    df = pd.read_excel(planilha)
                    protocolo = f"{df['Protocolo Glosa'][0]}".replace(".0", "")
                    self.driver.find_element(*self.codigo).send_keys(protocolo)
                    time.sleep(2)
                    self.driver.find_element(*self.pesquisar).click()
                    time.sleep(2)
                    self.driver.find_element(*self.recurso_de_glosa).click()
                    time.sleep(2)
                    content = self.driver.find_element(*self.body).text
                    recurso_iniciado = False

                    if 'Não existe informação de pagamento para a fatura recursada.' in content:
                        planilha_anterior = planilha
                        sem_extensao = planilha.replace('.xlsx', '')
                        novo_nome = sem_extensao + '_Sem_Pagamento.xlsx'
                        try:
                            time.sleep(2)
                            os.rename(planilha_anterior, novo_nome)
                        except PermissionError as err:
                            print(err)
                            continue

                    if 'A fatura não possui itens para gerar o lote de recurso de glosa ou já existem lotes gerados para a mesma.' in content:
                        recurso_iniciado = True
                        self.driver.find_element(*self.close_warning).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recurso_de_glosa_menu).click()
                        time.sleep(2)
                        self.driver.find_element(*self.fatura_input).send_keys(protocolo)
                        time.sleep(2)
                        self.driver.find_element(*self.pesquisar_recurso).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recurso_de_glosa_2).click()

                    self.inicializar_atributos(recurso_iniciado)
                    guias_abertas = False

                    pagina = 1
                    primeira_de_inicio = True
                    for index, linha in df.iterrows():
                        if f"{linha['Recursado no Portal']}" == "Sim" or f"{linha['Recursado no Portal']}" == "Não":
                            continue
                        nome_paciente = f'{linha["Paciente"]}'
                        numero_guia = f'{linha["Nro. Guia"]}'.replace('.0', '')
                        codigo_procedimento = f'{linha["Procedimento"]}'.replace('.0', '')
                        valor_glosa = f'{linha["Valor Glosa"]}'.replace('-', '').replace('.', ',')
                        valor_recurso = f'{linha["Valor Recursado"]}'
                        justificativa = f'{linha["Recurso Glosa"]}'.replace('\t', ' ')
                        # tabela_convenio = f'{linha["Tabela"]}'
                        matricula = f'{linha["Matrícula"]}'
                        pagina_iniciada = pagina
                        recurso = True
                        print(f'Paciente {nome_paciente}, N°Guia {numero_guia}, Código procedimento {codigo_procedimento}, Valor glosa: {valor_glosa}')

                        while recurso:
                            if guias_abertas == False:
                                table = self.driver.find_element(*self.table).get_attribute('outerHTML')
                                df_tabela = pd.read_html(table)[0]
                                time.sleep(1)

                                for i in range(1, len(df_tabela) + 1):
                                    try:
                                        self.driver.implicitly_wait(0.5)
                                        if recurso_iniciado == False:
                                            self.driver.find_element(By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[1]/i').click()
                                        else:
                                            self.driver.find_element(By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/table/tbody/tr[{i}]/td[1]/i').click()
                                    except Exception as e:
                                        continue

                                self.driver.implicitly_wait(15)
                                guias_abertas = True
                                vet_texto_label = self.driver.find_element(*self.label_registros).text.split(' ')
                                total_registros = int(vet_texto_label[9])
                                extensao_maxima_da_pagina = int(vet_texto_label[4])

                            for i in range(1, len(df_tabela) + 1):
                                nro_guia_portal, codigo_proc_portal, valor_glosa_portal, valor_recursado_portal, nome_paciente_portal = self.get_values(i, recurso_iniciado)
                                validacao_paciente = nome_paciente in nome_paciente_portal
                                validacao_matricula = matricula in nome_paciente_portal
                                validacao_numero_guia = numero_guia in nro_guia_portal
                                validacao_valor_glosa = valor_glosa in valor_glosa_portal
                                validacao_valor_recursado = valor_recursado_portal == "R$0,00"
                                validacao_codigo = codigo_procedimento in codigo_proc_portal
                                # validacao_codigo_taxa = "TAXAS" in tabela_convenio and "Taxa" in codigo_proc_portal


                                # validacao_numero_guia_alterado = validacao_paciente and not validacao_numero_guia and validacao_valor_glosa and validacao_valor_recursado and (validacao_codigo or validacao_codigo_taxa)
                        
                                # if validacao_numero_guia_alterado:
                                #     self.driver.switch_to.window(self.driver.window_handles[-1])
                                #     time.sleep(2)
                                #     numero_anterior = numero_guia
                                #     numero_guia = self.confere_numero_alterado(numero_guia, nro_guia_portal)
                                #     numero_alterado = numero_guia
                                #     if numero_anterior != numero_alterado:
                                #         df['Nro. Guia'] = df['Nro. Guia'].replace(int(numero_anterior), int(numero_alterado))
                                #         print(f'Paciente {nome_paciente}, N°Guia {numero_guia}, Código procedimento {codigo_procedimento}, Valor glosa: {valor_glosa}')
                                #     self.driver.switch_to.window(self.driver.window_handles[0])
                                    
                                validacao_normal = (validacao_numero_guia or validacao_paciente or validacao_matricula) and validacao_codigo and validacao_valor_glosa and validacao_valor_recursado
                                # validacao_taxa = (validacao_numero_guia or validacao_paciente or validacao_matricula) and validacao_codigo_taxa and validacao_valor_glosa and validacao_valor_recursado

                                if validacao_normal: #or validacao_taxa:
                                    input_valor_recursado, preencher_justificativa = self.xpath_preencher_valores(i, recurso_iniciado)
                                    self.driver.find_element(*input_valor_recursado).send_keys(valor_recurso)
                                    time.sleep(2)
                                    self.driver.find_element(*preencher_justificativa).click()
                                    time.sleep(2)
                                    for i in range(0, 10):
                                        try:
                                            self.driver.find_element(*self.text_area_justificativa).send_keys(justificativa)
                                            break
                                        except:
                                            time.sleep(2)
                                            continue
                                    time.sleep(2)
                                    self.driver.find_element(*self.button_ok).click()
                                    time.sleep(2)
                                    self.driver.find_element(*self.salvar_parcialmente).click()
                                    time.sleep(2)
                                    self.driver.find_element(*self.i_close).click()
                                    dados = {"Recursado no Portal" : ['Sim']}
                                    df_dados = pd.DataFrame(dados)
                                    book = load_workbook(planilha)
                                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                                    writer.book = book
                                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                                    df_dados.to_excel(writer, 'Recurso', startrow=index + 1, startcol=20, header=False, index=False)
                                    writer.save()
                                    recurso = False
                                    break                         
                            
                            if recurso == True:
                                if extensao_maxima_da_pagina == total_registros:
                                    self.driver.find_element(*self.primeira_pagina).click()
                                    guias_abertas = False
                                    pagina = 1
                                    time.sleep(2)
                                    if pagina == pagina_iniciada and primeira_de_inicio == False:
                                        primeira_de_inicio = True
                                        dados = {"Recursado no Portal" : ['Não']}
                                        df_dados = pd.DataFrame(dados)
                                        book = load_workbook(planilha)
                                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                                        writer.book = book
                                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                                        df_dados.to_excel(writer, 'Recurso', startrow=index + 1, startcol=20, header=False, index=False)
                                        writer.save()
                                        break
                                else:
                                    texto = self.driver.find_element(*self.ul).text
                                    vet_ul = texto.split('\n')
                                    if recurso_iniciado:
                                        primeira_de_inicio = False
                                        proxima_pagina = (By.XPATH, f'/html/body/main/div[1]/div[3]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/nav/ul/li[{len(vet_ul) - 1}]/a/span')
                                        pagina += 1
                                    else:
                                        primeira_de_inicio = False
                                        proxima_pagina = (By.XPATH, f'/html/body/main/div/div[2]/div/div/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/nav/ul/li[{len(vet_ul) - 1}]/a/span')
                                        pagina += 1
                                    contador_proxima_pagina = 0
                                    while contador_proxima_pagina < 10:
                                        try:
                                            self.driver.find_element(*proxima_pagina).click()
                                            break
                                        except:
                                            time.sleep(2)
                                            contador_proxima_pagina += 1
                                    time.sleep(2)
                                    guias_abertas = False
                    contador = 0
                    while contador <= 15:
                        try:
                            self.driver.find_element(*self.fechar).click()
                            break
                        except:
                            contador += 1
                            time.sleep(2)

                    time.sleep(2)        
                    self.driver.get("https://facpres.stm.jus.br/GuiasTISS/Relatorios/ViewRelatorioServicos")
                    time.sleep(2)
                    try:
                        writer.close()
                    except:
                        print("Erro ao fechar o writer")
                    planilha_anterior = planilha
                    sem_extensao = planilha.replace('.xlsx', '')
                    novo_nome = sem_extensao + '_Enviado.xlsx'
                    try:
                        time.sleep(2)
                        os.rename(planilha_anterior, novo_nome)
                    except PermissionError as err:
                        print(err)
                break
            except Exception as e:
                self.driver.get("https://facpres.stm.jus.br/GuiasTISS/Relatorios/ViewRelatorioServicos")
                count += 1
                print(e)
                content = self.driver.find_element(*self.body).text
                if 'LOGON' in content:
                    self.exe_login()
                    self.exe_caminho()
                    time.sleep(2)
                    self.driver.get("https://facpres.stm.jus.br/GuiasTISS/Relatorios/ViewRelatorioServicos")

    def confere_numero_alterado(self, numero_guia, nro_guia_portal):
        count = 0
        while count < 10:
            try:
                pesquisa = False
                while pesquisa == False:
                    try:
                        self.driver.find_element(*self.guia_op).clear()
                        time.sleep(1)
                        self.driver.find_element(*self.guia_op).send_keys(nro_guia_portal)
                        time.sleep(1)
                        self.driver.find_element(*self.buscar).click()
                        time.sleep(1)
                        pesquisa = True
                    except:
                        pass

                    user = False
                    while user == False:
                        try:
                            self.driver.find_element(By.XPATH, '//*[@id="menu_78B1E34CFC8E414D8EB4F83B534E4FB4"]').click()
                            self.driver.find_element(By.XPATH, '//*[@id="menu_78B1E34CFC8E414D8EB4F83B534E4FB4"]').click()
                            user = True
                        except:
                            pass
                senha = self.driver.find_element(By.XPATH, '/html/body/main/div[1]/div[2]/div/div[2]/div/div[2]/div[1]/div/div/div/div[2]/div[2]/div/div[2]/div[1]/div[3]/span').text
                if numero_guia in senha:
                    return nro_guia_portal
                
                else:
                    return numero_guia

            except:
                self.driver.refresh()
                count += 1
                if count == 10:
                    return numero_guia