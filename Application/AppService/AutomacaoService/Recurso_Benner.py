from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import load_workbook
import pandas as pd
import time
import os
from Application.AppService.AutomacaoService.page_element import PageElement


class inserir_dados_benner(PageElement):
    email = (By.XPATH, '//*[@id="Email"]')
    senha = (By.XPATH, '//*[@id="Senha"]')
    logar = (By.XPATH, '//*[@id="btnLogin"]')
    lotes_de_pagamento = (By.XPATH, '/html/body/div[3]/div[1]/div/ul/li[26]/a/span[1]')
    pesquisar_lotes = (By.XPATH, '/html/body/div[3]/div[1]/div/ul/li[26]/ul/li[3]/a/span')
    proximo = (By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[2]')
    fechar = (By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[3]')
    fechar_botao = (By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[2]')
    fechar_alerta = (By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')
    filtro = (By.XPATH, '//*[@id="Filtrar"]/span')
    protocolo = (By.XPATH, '//*[@id="Protocolo"]')
    pesquisar = (By.XPATH, '//*[@id="btn-Pesquisar"]/span')
    selecionar = (By.XPATH, '/html/body/div[3]/div[2]/div/div[2]/div/div/bc-smart-table-manager/div/div/div[2]/div/bc-smart-table/div[2]/table/tbody[1]/tr[1]/td[2]/input')
    atualizar = (By.XPATH, '//*[@id="atualizar-situacao-lote"]')
    recursoclick = (By.XPATH, '//*[@id="recurso-glosa"]/span')
    controle = (By.XPATH, '//*[@id="GuiasGlosadasTable"]/thead/tr[1]/th/div[2]/input')
    alerta2 = (By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[3]')
    marcar = (By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]/input')
    procedimento = (By.XPATH, '//*[@id="bcRecursar"]/span')
    checkbox = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[1]/input')
    checkbox2 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[1]/input')
    checkbox3 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[1]/input')
    checkbox4 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[1]/input')
    checkbox5 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[1]/input')
    checkbox6 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[1]/input')
    checkbox7 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[1]/input')
    checkbox8 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[1]/input')
    checkbox9 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[1]/input')
    checkbox10 = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[1]/input')
    recursar = (By.LINK_TEXT, 'Recursar Procedimento')
    justificativa = (By.XPATH, '/html/body/bc-modal-evolution/bc-modal-justificar-glosa/div/div/div/div[2]/div[2]/bc-exibir-a-partir-da-versao-tiss[2]/div/span/div/div/input')
    valor = (By.XPATH, '//*[@id="ValorRecursado"]')
    novo_recurso = (By.XPATH, '//*[@id="simpleConfirmationModal_btOk"]')
    pesquisar_proc = (By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/thead/tr[1]/th/div[2]/input')
    salvar = (By.XPATH, '/html/body/bc-modal-evolution/bc-modal-justificar-glosa/div/div/div/div[3]/button[2]')
    fechar_recurso = (By.XPATH, '/html/body/bc-modal-evolution/bc-modal-justificar-glosa/div/div/div/div[1]/button')
    fechar_botao = (By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[2]')
    fechar_alerta = (By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')
    filtro = (By.XPATH, '//*[@id="Filtrar"]/span')

    def __init__(self, url, usuario, senha):
        super().__init__(url)
        self.usuario = usuario
        self.senha = senha

    def exe_login(self):
        self.driver.find_element(*self.email).send_keys(self.usuario)
        self.driver.find_element(*self.senha).send_keys(self.senha)
        self.driver.find_element(*self.logar).click()

    def exe_caminho(self):
        try:
            modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div')))
            self.driver.find_element(*self.fechar_botao).click()

            while True:
                try:
                    proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[2]')))
                    proximo_botao.click()
                except:
                    break

            try:
                fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                fechar_botao.click()
            except:
                print("Não foi possível encontrar o botão de fechar.")
                pass

        except:
            print("Não teve Modal")
            pass
        self.driver.find_element(*self.lotes_de_pagamento).click()
        time.sleep(2)
        self.driver.find_element(*self.pesquisar_lotes).click()
        time.sleep(2)
        
        try:
            modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div')))
            self.driver.find_element(*self.fechar_botao).click()

            while True:
                try:
                    proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[2]')))
                    proximo_botao.click()
                except:
                    break

            try:
                fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                fechar_botao.click()
            except:
                print("Não foi possível encontrar o botão de fechar.")
                pass
        except:
            print("Não teve Modal")
            pass

        self.driver.find_element(*self.filtro).click()
        time.sleep(2)

    def Alert(self):
        self.driver.find_element(*self.proximo).click()
        time.sleep(1)
        self.driver.find_element(*self.fechar).click()                 

    def inicia_automacao(self, **kwargs):
        self.init_driver()
        self.open()
        self.exe_login()
        self.exe_caminho()

        diretorio = kwargs.get('diretorio')

        nomesarquivos = os.listdir(diretorio)

        for nome in nomesarquivos:
            self.driver.implicitly_wait(30)

            if "Enviado" in nome:
                print("PEG já enviado")
                continue

            sem_extensao = nome.replace('.xlsx', '')
            planilha = os.path.join(diretorio, nome)
            faturas_df1 = pd.read_excel(planilha)

            for index, linha in faturas_df1.iterrows():
                try:
                    self.driver.implicitly_wait(3)
                    modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div')))
                    self.driver.find_element(*self.fechar_botao).click()

                    while True:
                        try:
                            proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[2]')))
                            proximo_botao.click()
                        except:
                            break

                    try:
                        fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                        fechar_botao.click()
                    except:
                        print("Não foi possível encontrar o botão de fechar.")
                        pass
                except:
                    print("Não teve Modal")
                    pass
                self.driver.implicitly_wait(30)
                self.driver.find_element(*self.protocolo).send_keys(f"{linha['Protocolo Glosa']}")
                print('Protocolo: ' + f"{linha['Protocolo Glosa']}")
        
                time.sleep(1)
                self.driver.find_element(*self.pesquisar).click()
                time.sleep(2)
                try:
                    for contador in range(3):
                        self.driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
                except:
                    None

                try:
                    self.driver.implicitly_wait(3)
                    modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div')))
                    self.driver.find_element(*self.fechar_botao).click()

                    while True:
                        try:
                            proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[2]')))
                            proximo_botao.click()
                        except:
                            break

                    try:
                        fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                        fechar_botao.click()
                    except:
                        print("Não foi possível encontrar o botão de fechar.")
                        pass
                except:
                    print("Não teve Modal")
                    pass
                self.driver.implicitly_wait(30)

                self.driver.find_element(*self.selecionar).click()
                
                time.sleep(2)
                self.driver.find_element(*self.recursoclick).click()
                time.sleep(6)

                try:
                    self.driver.implicitly_wait(3)
                    element = WebDriverWait(self.driver, 50).until(EC.presence_of_element_located((self.novo_recurso)))
                    self.driver.find_element(*self.novo_recurso).click()
                except:
                    print('Recurso Normal')

                self.driver.implicitly_wait(30)
                time.sleep(2)
                try:
                    element = WebDriverWait(self.driver, 50).until(EC.presence_of_element_located((self.controle)))
                    self.driver.find_element(*self.controle).clear()
                except:
                    print('Limpar Nº Guia')

                break

            try:
                self.driver.implicitly_wait(3)
                modal = WebDriverWait(self.driver, 3.0).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div')))
                self.driver.find_element(*self.fechar_botao).click()

                while True:
                    try:
                        proximo_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="bcInformativosModal"]/div/div/div[3]/button[2]')))
                        proximo_botao.click()
                    except:
                        break

                try:
                    fechar_botao = WebDriverWait(self.driver, 0.2).until(EC.presence_of_element_located((By.XPATH, '/html/body/bc-modal-evolution/div/div/div/div[3]/button[3]')))
                    fechar_botao.click()
                except:
                    print("Não foi possível encontrar o botão de fechar.")
                    pass

            except:
                print("Não teve Modal")
                pass
            self.driver.implicitly_wait(30)

            count = 0
            faturas_df1 = pd.read_excel(planilha)

            for index, linha in faturas_df1.iterrows():
                self.driver.find_element(*self.controle).clear()
                time.sleep(1)
                self.driver.find_element(*self.controle).send_keys(f"{linha['Controle Inicial']}")
                time.sleep(1)
                self.driver.find_element(*self.marcar).click()
                time.sleep(1)
                self.driver.find_element(*self.procedimento).click()
                time.sleep(2)
                self.driver.find_element(*self.pesquisar_proc).clear()
                time.sleep(2) 
                self.driver.find_element(*self.pesquisar_proc).send_keys(f"{linha['Procedimento']}")
                time.sleep(2)
                self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/thead/tr[2]/th[8]').click()
                time.sleep(2)
                self.driver.find_element(*self.pesquisar_proc).clear()
                time.sleep(2)
                
                desmarcar = WebDriverWait(self.driver, 50).until(EC.presence_of_element_located((self.marcar)))
                time.sleep(1)
                self.driver.find_element(*self.controle).clear()
                time.sleep(1)
                self.driver.find_element(*self.controle).send_keys(f"{linha['Controle Inicial']}")
                self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]/input').click()
                time.sleep(2)
                self.driver.find_element(*self.controle).clear()
                print( 'Procedimentos ordenados')
                break

            for index, linha in faturas_df1.iterrows():
                print('---------------------------------------------------------------------------------------------------------')
                count = count + 1
                print(count)
                print('Controle: ' + f"{linha['Controle Inicial']}")

                if (f"{linha['Recursado no Portal']}") == "Sim":
                    print('Data: ' + f"{linha['Realizado']}")
                    print('Código: ' + f"{linha['Procedimento']}")
                    print('Procedimento já foi recursado no portal')
                    continue

                print('Data: ' + f"{linha['Realizado']}")
                time.sleep(2)
                self.driver.implicitly_wait(30)
                self.driver.find_element(*self.controle).clear()
                self.driver.find_element(*self.controle).send_keys(f"{linha['Controle Inicial']}")
                
                time.sleep(2)
                self.driver.find_element(*self.marcar).click()

                time.sleep(1)
                self.driver.find_element(*self.procedimento).click()
                time.sleep(2)
                self.driver.find_element(*self.pesquisar_proc).clear()
                time.sleep(2) 
                self.driver.find_element(*self.pesquisar_proc).send_keys(f"{linha['Procedimento']}")
                print('Código: ' + f"{linha['Procedimento']}")
                
                time.sleep(3)
                
                try:
                    for i in range(0,10):
                        try:
                            valor1 =  self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[8]').text.replace('R$ ', '')
                            codigo1 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[2]').text
                            for k, v in enumerate(codigo1):
                                if codigo1[0] != "0":
                                    break
                                if codigo1[k] != "0":
                                    codigo1 = str(codigo1[k:])
                                    break
                            vlr_glosado1 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[1]/tr[1]/td[7]').text.replace('R$ ', '')
                            elementos_encontrados = True
                            print("Elementos Encontrados")
                            break
                        except:
                            time.sleep(2)

                    valor_planilha = f"{linha['Valor Glosa']}".replace('R$', '')
                    procedimento_plan = f"{linha['Procedimento']}"

                    if ((valor1 == "0,00") and (codigo1 == procedimento_plan)) and (vlr_glosado1 == valor_planilha):
                        print('Valor do 1° Procedimento Encontrado = R$0,00')
                        print('Valor do 1° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)

                        self.driver.find_element(*self.checkbox).click()
                        time.sleep(1)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(1)
                        try:
                            button = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            time.sleep(1)
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            for i in range(10):
                                try:
                                    self.driver.find_element(*self.justificativa).click()
                                    time.sleep(1)
                                    self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                                    break
                                except:
                                    print('Não justificou')
                                    self.driver.find_element(*self.justificativa).clear()
                                    if i == 9:
                                        self.driver.quit()

                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(2)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(2)
                            print('Exceção')
                            continue

                    valor2 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo2 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo2):
                        if codigo2[0] != "0":
                            break
                        if codigo2[k] != "0":
                            codigo2 = str(codigo2[k:])
                            break
                    vlr_glosado2 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[2]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor2 == "R$0,00") and (codigo2 == f"{linha['Procedimento']}",".0")) and vlr_glosado2 == valor_planilha:
                        print('Valor do 2° Procedimento Encontrado = R$0,00')
                        print('Valor do 2° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox2).click()
                        time.sleep(1)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox2).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox2).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor3 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo3 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo3):
                        if codigo3[0] != "0":
                            break
                        if codigo3[k] != "0":
                            codigo3 = str(codigo3[k:])
                            break
                    vlr_glosado3 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[3]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor3 == "R$0,00") and (codigo3 == f"{linha['Procedimento']}",".0")) and vlr_glosado3 == valor_planilha:
                        print('Valor do 3° Procedimento Encontrado = R$0,00')
                        print('Valor do 3° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox3).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']
                            
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox3).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox3).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor4 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo4 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo4):
                        if codigo4[0] != "0":
                            break
                        if codigo4[k] != "0":
                            codigo4 = str(codigo4[k:])
                            break
                    vlr_glosado4 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[4]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor4 == "R$0,00") and (codigo4 == f"{linha['Procedimento']}",".0")) and vlr_glosado4 == valor_planilha:
                        print('Valor do 4° Procedimento Encontrado = R$0,00')
                        print('Valor do 4° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox4).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox4).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox4).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor5 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo5 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo5):
                        if codigo5[0] != "0":
                            break
                        if codigo5[k] != "0":
                            codigo5 = str(codigo5[k:])
                            break
                    vlr_glosado5 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[5]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor5 == "R$0,00") and (codigo5 == f"{linha['Procedimento']}",".0")) and vlr_glosado5 == valor_planilha:
                        print('Valor do 5° Procedimento Encontrado = R$0,00')
                        print('Valor do 5° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox5).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox5).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox5).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor6 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo6 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo6):
                        if codigo6[0] != "0":
                            break
                        if codigo6[k] != "0":
                            codigo6 = str(codigo6[k:])
                            break
                    vlr_glosado6 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[6]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor6 == "R$0,00") and (codigo6 == f"{linha['Procedimento']}",".0")) and vlr_glosado6 == valor_planilha:
                        print('Valor do 6° Procedimento Encontrado = R$0,00')
                        print('Valor do 6° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox6).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            time.sleep(2)
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox6).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox6).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                        

                    valor7 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo7 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo7):
                        if codigo7[0] != "0":
                            break
                        if codigo7[k] != "0":
                            codigo7 = str(codigo7[k:])
                            break
                    vlr_glosado7 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[7]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor7 == "R$0,00") and (codigo7 == f"{linha['Procedimento']}",".0")) and vlr_glosado7 == valor_planilha:
                        print('Valor do 7° Procedimento Encontrado = R$0,00')
                        print('Valor do 7° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox7).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox7).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox7).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor8 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo8 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo8):
                        if codigo8[0] != "0":
                            break
                        if codigo8[k] != "0":
                            codigo8 = str(codigo8[k:])
                            break
                    vlr_glosado8 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[8]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor8 == "R$0,00") and (codigo8 == f"{linha['Procedimento']}",".0")) and vlr_glosado8 == valor_planilha:
                        print('Valor do 8° Procedimento Encontrado = R$0,00')
                        print('Valor do 8° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox8).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox8).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox8).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor9 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo9 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo9):
                        if codigo9[0] != "0":
                            break
                        if codigo9[k] != "0":
                            codigo9 = str(codigo9[k:])
                            break
                    vlr_glosado9 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[9]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor9 == "R$0,00") and (codigo9 == f"{linha['Procedimento']}",".0")) and vlr_glosado9 == valor_planilha:
                        print('Valor do 9° Procedimento Encontrado = R$0,00')
                        print('Valor do 9° Procedimento injetado = ' + f"{linha['Valor Recursado']}")
                        time.sleep(1)
                        self.driver.find_element(*self.checkbox9).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox9).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox9).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue

                    valor10 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[8]').text.replace('R$ ', '')
                    codigo10 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[2]').text
                    for k, v in enumerate(codigo10):
                        if codigo10[0] != "0":
                            break
                        if codigo10[k] != "0":
                            codigo10 = str(codigo10[k:])
                            break
                    vlr_glosado10 = self.driver.find_element(By.XPATH, '//*[@id="ProcedimentosGlosadosTable"]/tbody[10]/tr[1]/td[7]').text.replace('R$ ', '')

                    if ((valor10 == "R$0,00") and (codigo10 == f"{linha['Procedimento']}",".0")) and vlr_glosado10 == valor_planilha:
                        print('Valor do 10° Procedimento Encontrado = R$0,00')
                        print('Valor do 10° Procedimento injetado = ' + f"{linha['Valor Recursado']}")

                        time.sleep(1)
                        self.driver.find_element(*self.checkbox10).click()
                        time.sleep(2)
                        self.driver.find_element(*self.recursar).click()
                        time.sleep(2)
                        try:
                            button= WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="ValorRecursado"]')))
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)
                            dados = ['Sim']
                            df = pd.DataFrame(dados)
                            
                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow= count, startcol=20, header=False, index=False)

                            writer.save()
                        except:
                            print('Erro Valor')
                            self.driver.find_element(*self.valor).click()
                            print('clicou no valor')
                            time.sleep(1)
                            self.driver.find_element(*self.valor).send_keys(f"{linha['Valor Recursado']}")
                            print('Valor inserido')
                            time.sleep(1)
                            self.driver.find_element(*self.justificativa).send_keys(f"{linha['Recurso Glosa']}")
                            print('Justificativa inserida')
                            time.sleep(2)
                            self.driver.find_element(*self.salvar).click()
                            print('salvo')
                            time.sleep(1)

                            dados = ['Sim']

                            df = pd.DataFrame(dados)

                            book = load_workbook(planilha)

                            #defino o writer para escrever em um novo arquivo 
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')

                            #incluo a formatação no writer
                            writer.book = book

                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                            # injetar os valores em um lugar específico:
                            df.to_excel(writer, 'Recurso', startrow = count, startcol=20, header=False, index=False)

                            writer.save()
                        try:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox10).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            continue
                        except:
                            time.sleep(2)
                            self.driver.find_element(*self.checkbox10).click()
                            time.sleep(1)
                            self.driver.find_element(*self.controle).clear()
                            time.sleep(1)
                            self.driver.find_element(By.XPATH, '//*[@id="GuiasGlosadasTable"]/tbody[1]/tr[1]/td[1]').click()
                            time.sleep(3)
                            print('Exceção')
                            continue
                except:
                    print('Procedimento já recursados ou não existe esse código nesse protocolo.')
                    desmarcar = WebDriverWait(self.driver, 50).until(EC.presence_of_element_located((self.marcar))).click()
                    continue
            novo_nome = diretorio + '/' + sem_extensao + '_Enviado.xlsx'
            try:
                writer.close()
                book.close()
            except:
                pass
            try:
                os.rename(planilha, novo_nome)
            except:
                print("Erro ao renomear arquivo")
            self.driver.get('https://portalconectasaude.com.br/Pagamentos/PesquisaLote/PesquisaComValorGlosado')
            time.sleep(2)
            self.driver.find_element(*self.filtro).click()

        self.driver.quit()