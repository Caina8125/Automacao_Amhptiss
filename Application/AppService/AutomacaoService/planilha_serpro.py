from tkinter.messagebox import showerror, showinfo
from pandas import DataFrame
from pandas import ExcelWriter
from pandas import read_excel
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter.filedialog import askopenfilename
from openpyxl.worksheet.worksheet import Worksheet


class PlanilhaSerpro(): 
    def __init__(self, path_planilha: str) -> None:
        self.path_planilha: str = path_planilha
        self._df_planilha: DataFrame = read_excel(path_planilha)
        self._df_planilha['Valor Liberado'] = ''
        self._df_planilha = self._df_planilha.loc[:, [
            'Fatura', 'Protocolo Glosa', 'Valor Recursado', 'Amhptiss', 'Nro. Guia', 'Autorização', 'Matrícula', 'Paciente', 'Procedimento',
            'Descrição','Valor Original', 'Valor Liberado', 'Valor Glosa', 'Motivo Glosa', 'Valor Cobrado', 'Recurso Glosa'
            ]]
        self._book = None
        self.sheet = None
    
    def create_writer(self, numero_processo: str) -> ExcelWriter:
        self._book: Workbook = load_workbook(r"Templates\SERPRO.xlsx")
        self.sheet = self._book['Recurso de Glosas Serpro']
        writer: ExcelWriter = ExcelWriter(f'output\\SERPRO\\Serpro_{numero_processo}.xlsx')
        writer.book = self._book
        writer.sheets = dict((ws.title, ws) for ws in self._book.worksheets)
        return writer
    
    def filter_df_by_number(self, numero_processo) -> DataFrame:
        df_processo: DataFrame = self._df_planilha.loc[(self._df_planilha["Fatura"] == int(numero_processo))]

        if df_processo.empty:
            df_processo:DataFrame = self._df_planilha.loc[(self._df_planilha["Fatura"] == numero_processo)]

        for index, linha in df_processo.iterrows():
            valor_liberado = linha['Valor Original'] - abs(linha['Valor Glosa'])
            df_processo['Valor Liberado'][index] = valor_liberado
        
        return df_processo
    
    def get_info_processo(self, df_processo: DataFrame) -> None:
        return {
            'protocolo': f'{df_processo["Protocolo Glosa"].values.tolist()[0]}'.replace('.0', ''),
            'numero_fatura': f'{df_processo["Fatura"].values.tolist()[0]}',
            'valor_total_original': df_processo["Valor Original"].sum(),
            'valor_liberado': df_processo["Valor Original"].sum() - abs(df_processo["Valor Glosa"].sum()),
            'valor_glosa_total': df_processo["Valor Glosa"].sum(),
            'valor_recurso_total': df_processo["Valor Recursado"].sum()
        }
    
    def atualiza_template(self, chave, dado) -> None:
        match chave:
            case 'protocolo':
                self.add_info_na_celula('Protocolo:', dado)

            case 'numero_fatura':
                self.add_info_na_celula('Lote:', dado)

            case 'valor_total_original':
                self.add_info_na_celula('Valor Informado:', f'R${dado:_.2f}'.replace('.',',').replace('_','.'))

            case 'valor_liberado':
                self.add_info_na_celula('Valor Liberado:', f'R${dado:_.2f}'.replace('.',',').replace('_','.'))

            case 'valor_glosa_total':
                self.add_info_na_celula('Valor Glosa:', f'R${dado:_.2f}'.replace('.',',').replace('_','.'))

            case 'valor_recurso_total':
                self.add_info_na_celula('Valor Recurso:', f'R${dado:_.2f}'.replace('.',',').replace('_','.'))

    def valor_in_cell(self, valor: str, cell: str | None) -> bool:
        try:
            if valor in cell:
                return True
            else:
                return False
        except:
            return False
    
    def add_info_na_celula(self, valor_celula: str, valor: str) -> None:
        celula_encontrada = False

        for row in self.sheet.iter_rows():

            for cell in row:
                
                if self.valor_in_cell(valor_celula, cell.value):
                    new_value = valor_celula + ' ' + valor
                    self.sheet.cell(row=cell.row, column=cell.column).value = new_value
                    celula_encontrada = True
                    break

            if celula_encontrada:
                break
            
    
    def create_excel(self, writer: ExcelWriter, df_processo: DataFrame) -> None:
        df_processo = df_processo.drop(['Fatura', 'Protocolo Glosa', 'Valor Recursado'], axis='columns')
        df_processo.to_excel(writer, 'Recurso de Glosas Serpro', startrow=5, startcol=0, header=False, index=False)

    def gerar_arquivos_excel(self) -> None:
        LISTA_DE_PROCESSOS: list[str] = [
            f'{value}'.replace('.0', '') 
            for value in list(set(self._df_planilha['Fatura'].values.tolist()))
            ]

        for numero_processo in LISTA_DE_PROCESSOS:
            df_processo: DataFrame = self.filter_df_by_number(numero_processo)
            info_processo: dict = self.get_info_processo(df_processo)
            writer: ExcelWriter = self.create_writer(numero_processo)

            for chave, dado in info_processo.items():
                self.atualiza_template(chave, dado)

            self.create_excel(writer, df_processo)
            writer.save()
        
def exec_planilha():
    try:
        path_planilha = askopenfilename()
        teste = PlanilhaSerpro(path_planilha)
        teste.gerar_arquivos_excel()
        showinfo('Automação', 'Planilhas geradas com sucesso!')
    except Exception as e:
        showerror('Automação', f'Ocorreu uma exceção não tratada.\n{e.__class__.__name__}:\n{e}')