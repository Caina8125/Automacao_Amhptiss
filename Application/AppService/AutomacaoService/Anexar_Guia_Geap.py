import pandas as pd
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
from Application.AppService.AutomacaoService.page_element import PageElement
    
class AnexarGuiaGeap(PageElement):
    acessar_portal = (By.XPATH, '/html/body/div[3]/div[3]/div[1]/form/div[1]/div[1]/div/a')
    usuario = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/div/label[1]/div/div[1]/div/input')
    senha = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/div/label[2]/div/div[1]/div[1]/input')
    entrar = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[2]/button/div[2]/div/div')
    fechar = (By.XPATH, '/html/body/div[4]/div[2]/div/div[3]/button')
    portal_tiss = (By.XPATH, '/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div[1]/div/div')
    versao_anterior = (By.XPATH, '/html/body/div[1]/div/div[1]/aside/div[1]/div[3]/button/span[2]/span')
    alerta = (By.XPATH,' /html/body/div[2]/div/center/a')
    guia = (By.XPATH,'//*[@id="objTableDetalhe"]/tbody/tr[3]/td[1]/a')
    envio_xml = (By.XPATH,'//*[@id="main"]/div/div/div[2]/div/article/div[6]/div[4]/div[4]/div[4]/div/div[2]/ul/li[2]/a')
    portal_tiss = (By.XPATH, '/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div[1]/div/div')
    anexar = (By.XPATH,'//*[@id="fupDoc"]')
    adicionar = (By.XPATH,'//*[@id="btnAdicionar"]')
    url_guia = 'https://wwwapi.geap.com.br/AuditoriaDigital/api/v1/guias/'
    quant_guia_consensuada = 0
    quant_guia_cancelada = 0
    lista_df = []
    numero_envio = (By.XPATH, '/html/body/form/table/tbody/tr[5]/td[2]/input')
    listar = (By.XPATH, '//*[@id="MenuOptionReport"]')
    body = (By.XPATH, '/html/body')
    sem_erros = (By.XPATH,'//*[@id="StaErro"]/option[2]')
    qtd_guias = (By.XPATH, '/html/body/form/table/tbody/tr[5]/td/table/tbody/tr[2]/td[7]/a')
    total = (By.XPATH, '/html/body/form/table/tbody/tr[5]/td/table/tbody/tr[1]/td/table/tbody/tr/td[6]')
    acessar_portal = (By.XPATH, '/html/body/div[3]/div[3]/div[1]/form/div[1]/div[1]/div/a')
    fechar = (By.XPATH, '/html/body/div[4]/div[2]/div/div[3]/button')
    arrow = (By.XPATH, '/html/body/div[1]/div/header/div[1]/div[2]/button/span[2]/span/i')
    sair = (By.XPATH, '/html/body/div[3]/div/div[3]/div[2]/div')
    confirmar = (By.XPATH, '/html/body/div[1]/div[2]/div/div/ui-panel/div[2]/button/div[2]/div/div')

    def __init__(self, url, cpf, senha):
        super().__init__()
        self.url = url
        self.cpf = cpf
        self.senha = senha

    def exe_login(self):
        time.sleep(4)
        try:
            self.driver.find_element(*self.fechar).click()
        except:
            pass
        self.driver.find_element(*self.acessar_portal).click()
        try:
            self.driver.implicitly_wait(4)
            self.driver.find_element(*self.usuario).send_keys(self.cpf)
            self.driver.find_element(*self.senha).click()
            self.driver.find_element(*self.senha).send_keys(self.senha)
            time.sleep(2)
            self.driver.find_element(*self.entrar).click()
            time.sleep(2)
            self.driver.find_element(*self.portal_tiss).click()
            
        except:
            self.driver.implicitly_wait(180)
            self.driver.find_element(*self.portal_tiss).click()
    
    def exe_caminho(self):
        self.driver.switch_to.window(self.driver.window_handles[-1])
        self.driver.implicitly_wait(4)
        try:
            self.driver.find_element(*self.alerta).click()
        except:
            print('Alerta não apareceu')
        self.driver.implicitly_wait(15)
        
        # self.driver.find_element(*self.versao_anterior).click()
        time.sleep(2)

        self.driver.find_element(*self.envio_xml).click()
        time.sleep(1)
        # driver.switch_to.window(driver.window_handles[1])
        self.driver.switch_to.window(self.driver.window_handles[-1])
    
    def inicia_automacao(self, planilha):
        self.open()
        self.exe_login()
        self.exe_caminho()
        for _ in range(0,10):
            try:
                count = 0
                faturas_df = pd.read_excel(planilha)
                numero_envio_anterior = ''

                for index, linha in faturas_df.iterrows():
                    numero_envio = str(linha['Nº Envio']).replace('.0', '')

                    if (f"{linha['Guia Anexada']}") == "Sim" or (f"{linha['Guia Anexada']}") == "Não Anexado" or (f"{linha['Guia Anexada']}") == "Consensuada" or (f"{linha['Guia Anexada']}") == "Cancelada":
                        print(f"{linha['Guia Anexada']}")
                        count = count + 1
                        continue

                    if numero_envio != numero_envio_anterior:
                        conteudo = self.driver.find_element(*self.body).text

                        if 'Relatório' in conteudo:
                            self.driver.close()
                            self.driver.switch_to.window(self.driver.window_handles[-1])
                            self.driver.back()

                        self.driver.find_element(*self.numero_envio).clear()
                        self.driver.find_element(*self.numero_envio).send_keys(numero_envio)
                        time.sleep(2)
                        # self.driver.find_element(*self.sem_erros).click()
                        # time.sleep(2)
                        self.driver.find_element(*self.listar).click()
                        content = self.driver.find_element(*self.body).text

                        if 'Não existem registros na base da dados para o critério escolhido.' in content:
                            self.driver.back()
                            continue

                        time.sleep(2)
                        self.driver.find_element(*self.qtd_guias).click()



                    else:
                        print('Guia pronta para ser anexada')

                    self.driver.switch_to.window(self.driver.window_handles[-1])
                    count = count + 1
                    time.sleep(2)
                    numero_geap = f"{linha['Nro Guia GEAP']}".replace('.0', '')
                    table = self.driver.find_element(By.XPATH, '//*[@id="objTableDetalhe"]')
                    tabela_html = table.get_attribute('outerHTML')
                    df_processo = pd.read_html(tabela_html, skiprows=2, header=0)[0]
                    df_processo = df_processo.iloc[:-1]
                    quantidade_de_guias = len(df_processo) + 3

                    for i in range(3, quantidade_de_guias):
                        id_portal = self.driver.find_element(By.XPATH, f'/html/body/form/table/tbody/tr[5]/td/table/tbody/tr[{i}]/td[1]/a').text

                        if id_portal == numero_geap:
                            try:
                                self.driver.find_element(By.XPATH, f'/html/body/form/table/tbody/tr[5]/td/table/tbody/tr[{i}]/td[1]/a').click()
                            except:
                                continue
                            break

                        else:
                            continue

                    print('Entrando na guia')
                    time.sleep(2)
                    content = self.driver.find_element(*self.body).text

                    if 'Essa Guia está com erro, não será possível anexar arquivo.' in content:
                        dados = ['Guia com erro']
                        df = pd.DataFrame(dados)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, 'Planilha1', startrow= count, startcol=4, header=False, index=False)
                        writer.save()
                        self.driver.back()
                        continue

                    self.driver.find_element(*self.anexar).send_keys(linha["Caminho"])
                    time.sleep(1)
                    self.driver.find_element(*self.adicionar).click()
                    time.sleep(2)

                    try:
                        id = self.driver.find_element(By.XPATH,'//*[@id="grvLista"]/tbody/tr[1]/th[1]').text

                    except:
                        print('Erro ao anexar guia')
                        dados = ['Não Anexado']
                        df = pd.DataFrame(dados)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, 'Planilha1', startrow= count, startcol=4, header=False, index=False)
                        writer.save()
                        self.driver.back()
                        self.driver.back()
                        numero_envio_anterior = numero_envio
                        continue

                    if id == "Id Arquivo":
                        print('Guia Anexada')                                                                   
                        dados = ['Sim']
                        df = pd.DataFrame(dados)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, 'Planilha1', startrow= count, startcol=4, header=False, index=False)
                        writer.save()

                    else:
                        print('Erro ao anexar guia')
                        dados = ['Cancelado']
                        df = pd.DataFrame(dados)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, 'Planilha1', startrow= count, startcol=4, header=False, index=False)
                        writer.save()

                    self.driver.back()
                    self.driver.back()
                    numero_envio_anterior = numero_envio
                break

            except Exception as e:
                handles = self.driver.window_handles
                numero_de_guias = len(handles)

                for g in range(1, numero_de_guias - 1):
                    self.driver.close()
                    self.driver.switch_to.window(self.driver.window_handles[-1])

                self.exe_caminho()
        