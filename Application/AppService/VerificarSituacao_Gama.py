import pandas as pd
import time
from tkinter import filedialog
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from Filtro_Faturamento import *
from selenium.webdriver.chrome.options import Options
from seleniumwire import webdriver
from page_element import PageElement

class Login(PageElement):
    usuario = (By.XPATH, '//*[@id="username"]')
    senha = (By.XPATH, '//*[@id="password"]')
    entrar = (By.XPATH, '//*[@id="submitPrestador"]')

    def exe_login(self, usuario, senha):
        self.driver.find_element(*self.usuario).send_keys(usuario)
        self.driver.find_element(*self.senha).send_keys(senha)
        self.driver.find_element(*self.entrar).click()

class Caminho(PageElement):
    gama = (By.XPATH, '//*[@id="TB_ajaxContent"]/div/div[1]/div/div/a/img')
    autorizacao = (By.XPATH, '//*[@id="header"]/div[2]/div/div[2]/ul/li[2]/a')
    status = (By.XPATH, '//*[@id="header"]/div[2]/div/div[2]/ul/li[2]/ul/li[3]/a')

    def exe_caminho(self):
        self.driver.find_element(*self.gama).click()
        time.sleep(2)
        self.driver.find_element(*self.autorizacao).click()
        self.driver.find_element(*self.status).click()
        time.sleep(2)

class FazerPesquisa(PageElement):
    senha = (By.XPATH, '//*[@id="txtCodigoSenha"]')
    consultar = (By.XPATH, '//*[@id="btnConsultarSenha"]')
    status = (By.XPATH, '//*[@id="dadosConsultarStatusAutorizacao"]/div[2]/fieldset/div[2]/div[2]/span')
    numero_da_carteira = (By.XPATH, '//*[@id="dadosConsultarStatusAutorizacao"]/div[3]/fieldset/div[1]/div/span')
    tabela_procedimento = (By.XPATH, '//*[@id="dadosConsultarStatusAutorizacao"]/div[5]/fieldset/table')

    def fazer_pesquisa(self):
        df = pd.read_excel(planilha)
        senha_loc = ""
        count = 0
        for index, linha in df.iterrows():
            senha = f"{linha['Senha Aut.']}"
            if senha == "0":
                count += 1
                continue
            if (f"{linha['Pesquisado no Portal']}") == "Sim":
                print('Já foi feita a pesquisa desta autorização.')
                count += 1
                print(count)
                print('___________________________________________________________________________')
                continue

            if senha_loc == senha:
                count += 1
                continue
            
            senha_loc = f"{linha['Senha Aut.']}"

            count += 1

            senha_df = df.loc[df["Senha Aut."] == senha]

            count2 = 0

            self.driver.find_element(*self.senha).send_keys(senha)
            self.driver.find_element(*self.consultar).click()
            time.sleep(2)

            for index2, linha2 in senha_df.iterrows():
                if (f"{linha2['Pesquisado no Portal']}") == "Sim":
                    print('Já foi feita a pesquisa desta autorização.')
                    print(count)
                    print('___________________________________________________________________________')
                    continue

                try:
                    status = self.driver.find_element(*self.status).text
                    matricula_portal = self.driver.find_element(*self.numero_da_carteira).text
                    tabela = self.driver.find_element(*self.tabela_procedimento)
                    tabela_html = tabela.get_attribute('outerHTML')
                    df_procedimento = pd.read_html(tabela_html)[0]
                    lista_procedimentos = df_procedimento.values.tolist()
                    procedimento_planilha = f"{linha2['Procedimento']}".replace('.', '').replace('-', '')
                    matricula_planilha = f"{linha2['Matríc. Convênio']}"

                    if str(matricula_planilha) == str(matricula_portal):
                        resposta_matricula = 'Válida'
                    else:
                        resposta_matricula = 'Inválida'

                    if procedimento_planilha[0] == '1' and len(procedimento_planilha) == 9:
                        data = {'Situação': [status], 'Validação Carteira': [resposta_matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Pesquisado no Portal': ['Sim']}
                        df_dados = pd.DataFrame(data)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df_dados.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                        writer.save()
                        count2 += 1
                        continue   

                    if procedimento_planilha[0] == "0" or procedimento_planilha[0] == "6" or procedimento_planilha[0] == "7" or procedimento_planilha[0] == "8":
                        data = {'Situação': [status], 'Validação Carteira': [resposta_matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Pesquisado no Portal': ['Sim']}
                        df_dados = pd.DataFrame(data)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df_dados.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                        writer.save()
                        count2 += 1
                        continue 

                    while procedimento_planilha[0] == "9":
                        if procedimento_planilha[0] == "9" and procedimento_planilha[1] == "8":
                            print("Procedimento")
                            break

                        else:
                            data = {'Situação': [status], 'Validação Carteira': [resposta_matricula], 'Validação Proc.': ['Mat/Med, Taxas'], 'Pesquisado no Portal': ['Sim']}
                            df_dados = pd.DataFrame(data)
                            book = load_workbook(planilha)
                            writer = pd.ExcelWriter(planilha, engine='openpyxl')
                            writer.book = book
                            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                            df_dados.to_excel(writer, "Sheet1", startrow= count + count2, startcol=7, header=False, index=False)
                            writer.save()
                            count2 += 1
                            break

                    if procedimento_planilha[0] == "9" and procedimento_planilha[1] != "8":
                        continue
                    
                    procedimento_encontrado = False
                    for lista in lista_procedimentos:
                        procedimento_portal = lista[0]
                        if int(procedimento_planilha) == int(procedimento_portal):
                            resposta_procedimento = 'Ok'
                            procedimento_encontrado = True
                            break
                    if procedimento_encontrado == False:
                        resposta_procedimento = 'Não consta nesta autorização'

                    dados = {"Situação" : [status], "Validação Carteira": [resposta_matricula], "Validação Proc.": [resposta_procedimento], "Pesquisado no Portal" : ["Sim"]}
                    df_dados = pd.DataFrame(dados)
                    book = load_workbook(planilha)
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df_dados.to_excel(writer, 'Sheet1', startrow= count + count2, startcol=7, header=False, index=False)
                    writer.save()
                    count2 += 1
                except:
                    status = "Senha não encontrada"
                    dados = {"Situação" : [status], "Validação Carteira": [""], "Validação Proc.": [""], "Pesquisado no Portal" : ["Sim"]}
                    df_dados = pd.DataFrame(dados)
                    book = load_workbook(planilha)
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df_dados.to_excel(writer, 'Sheet1', startrow= count + count2, startcol=7, header=False, index=False)
                    writer.save()
                    count2 += 1
            self.driver.get('https://wwwt.connectmed.com.br/conectividade/prestador/consulta/statusAutorizacao.htm')
#-------------------------------------------------------------------------------------------------

def verificar_gama(user, password):
    try:
        processar_planilha()
        remove()
    except:
        pass
    
    global planilha
    planilha = filedialog.askopenfilename()

    url = 'https://wwwt.connectmed.com.br/conectividade/prestador/home.htm'

    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_argument('--ignore-ssl-errors')

    options = {
    'proxy': {
            'http': f'http://{user}:{password}@10.0.0.230:3128',
                'https': f'http://{user}:{password}@10.0.0.230:3128'
        }
    }
    try:
        servico = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=servico, seleniumwire_options=options, options=chrome_options)
    except:
        driver = webdriver.Chrome(seleniumwire_options=options, options=chrome_options)

    login_page = Login(driver, url)
    login_page.open()
    login_page.exe_login(
        usuario = "AMHPFATURAMENTO",
        senha = "fat0073"
    )

    time.sleep(2)

    Caminho(driver, url).exe_caminho()
    FazerPesquisa(driver, url).fazer_pesquisa()        
