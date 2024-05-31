import tkinter.messagebox
from tkinter import filedialog
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium import webdriver
from abc import ABC
import pandas as pd
import time
import os
from selenium.webdriver.chrome.options import Options
from seleniumwire import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import Pidgin
from page_element import PageElement

class Login(PageElement):
    usuario = (By.XPATH, '//*[@id="cpfOuCnpj"]')
    senha = (By.XPATH, '//*[@id="senha"]')
    acessar = (By.XPATH, '//*[@id="loginGeral"]')

    def exe_login(self, usuario, senha):
        self.driver.implicitly_wait(30)
        time.sleep(1.5)
        self.driver.find_element(*self.usuario).send_keys(usuario)
        time.sleep(1.5)
        self.driver.find_element(*self.senha).send_keys(senha)
        time.sleep(1.5)
        self.driver.find_element(*self.acessar).click()
        time.sleep(1.5)

class caminho(PageElement):
    finalizar = (By.XPATH, '//*[@id="step-0"]/nav/button')
    demonstrativo_tiss = (By.XPATH, '/html/body/div[1]/aside/section/div/div/div[1]/div[1]/ul/li[11]/a')
    demonstrativo_de_analises = (By.XPATH, '/html/body/div[1]/aside/section/div/div/div[1]/div[1]/ul/li[11]/ul/li[1]/a')

    def exe_caminho(self):
        self.driver.implicitly_wait(30)
        self.driver.find_element(*self.finalizar).click()
        time.sleep(4)
        self.driver.get('https://servicosonline.cassi.com.br/Prestador/RecursoRevisaoPagamento/TISS/DemonstrativoAnaliseContas/Index')

class Recursar(PageElement):
    protocolo_input = (By.XPATH, '//*[@id="ProtocoloPagamento"]')
    consultar = (By.ID, 'btnConsultar')
    voltar = (By.XPATH, '//*[@id="btnVoltar"]')
    xpath_corpo_da_pagina = (By.XPATH, '/html/body')
    demonstrativo_de_analise = (By.XPATH, '/html/body/div[1]/div[5]/section/div/fieldset/div/table/tbody/tr/td[3]/form/input[3]')
    recursar = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[1]/div[1]/div/div[2]/button[2]')
    guia_input = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[7]/div[2]/div/div/fieldset/div[1]/div/input')
    pesquisar = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[7]/div[2]/div/div/fieldset/div[1]/div/div/a')
    guia_click = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[7]/div[2]/div/div/fieldset/div[2]/table/tbody/tr/td[1]')
    valor_recursando_input = (By.ID, 'txtValorRecusado')
    mostrar_textarea_just = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[3]/div/div/div[2]/div[1]/div[6]/div[2]/div/div/div[1]/h4/a')
    textarea_justificativa = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[3]/div/div/div[2]/div[1]/div[5]/div/div/div/div/textarea')
    fechar_justificativa = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[3]/div/div/div[2]/div[1]/div[5]/div/div/div/div/div/button')
    salvar_recurso = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[3]/div/div/div[2]/div[1]/div[8]/div/button[1]')
    fechar_recurso = (By.ID, 'btnFecharJustificativaEvento')
    ok = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[6]/div/div/div[3]/button')
    solicitar = (By.XPATH, '/html/body/div[1]/div[5]/section/nav/div/div[2]/ul[1]/li[3]/a')
    editar_recurso = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[10]/div/table/tbody/tr/td[10]/a[1]')
    proxima_pagina = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[10]/div/center/table/tbody/tr/td[3]/form/a')
    table = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[7]/div[3]/div[1]/div[2]/fieldset/div/table')
    protocolo_de_revisao_input = (By.XPATH, '/html/body/div[1]/div[5]/section/div/form/fieldset/div[2]/div[4]/input')
    resposta_recurso = (By.XPATH, '/html/body/div[1]/div[5]/section/div/fieldset/div/table/tbody/tr/td[4]/form/input[4]')
    contestar = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[1]/div[1]/div/div[2]/button[1]')
    acao = (By.ID, '#dropDownContestar')
    contestar_opt = (By.ID, 'btnContestarRecurso')
    div_lista_guias = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[7]/div[2]/div/div/fieldset/div[2]')
    protocolo_contestacao = (By.XPATH, '/html/body/div[1]/div[5]/section/div/div[7]/div[1]/div/table/tbody/tr/td[1]')

    def recurso(self, pasta):
        lista_de_planilhas = [f"{pasta}/{arquivo}" for arquivo in os.listdir(pasta) if arquivo.endswith('.xlsx')]
        for i in range(0,10):
            try:

                for planilha in lista_de_planilhas:
                    df = pd.read_excel(planilha)
                    protocolo = f"{df['Protocolo Aceite'][0]}".replace('.0', '').strip()

                    if "P" in protocolo:
                        protocolo_pagamento = f"{df['Protocolo Pagamento'][0]}".replace('.0', '')
                        
                    else:
                        protocolo_pagamento = ''

                    if "Enviado" in planilha:
                        continue

                    if "P" in protocolo and not protocolo_pagamento.isdigit():    
                        self.driver.get('https://servicosonline.cassi.com.br/Prestador/RecursoRevisaoPagamento/TISS/DemonstrativoRecursoGlosa/Index')
                        time.sleep(2)
                        self.driver.find_element(*self.protocolo_de_revisao_input).send_keys(protocolo)
                        time.sleep(2)
                        self.driver.find_element(*self.consultar).click()
                        time.sleep(2)
                        self.driver.find_element(*self.resposta_recurso).click()
                        time.sleep(2)
                        self.driver.find_element(*self.contestar).click()
                        time.sleep(2)
                        self.driver.find_element(*self.acao).click()
                        time.sleep(2)
                        self.driver.find_element(*self.contestar_opt).click()
                        time.sleep(2)
                        protocolo_contestacao = self.driver.find_element(*self.protocolo_contestacao).text
                        time.sleep(2)
                        dados = {"Protocolo": [protocolo_contestacao]}
                        df_dados = pd.DataFrame(dados)
                        book = load_workbook(planilha)
                        writer = pd.ExcelWriter(planilha, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df_dados.to_excel(writer, 'Recurso', startrow=1, startcol=1, header=False, index=False)
                        writer.save()
                        writer.close()
                        
                    else:    
                        self.driver.find_element(*self.protocolo_input).send_keys(protocolo)
                        time.sleep(2)
                        self.driver.find_element(*self.consultar).click()
                        time.sleep(2)
                        corpo_pagina = self.driver.find_element(*self.xpath_corpo_da_pagina).text
                        time.sleep(1)

                        if not "Não foram encontrados resultados para a pesquisa" in corpo_pagina:

                            if not 'Campo "Protocolo de Pagamento": Informe apenas números' in corpo_pagina:
                                self.driver.find_element(*self.demonstrativo_de_analise).click()
                                time.sleep(2)
                                self.driver.find_element(*self.recursar).click()
                                time.sleep(2)
                                corpo_pagina = self.driver.find_element(*self.xpath_corpo_da_pagina).text

                            if 'Ação não permitida. Já existe um recurso/revisão em digitação para este protocolo.' in corpo_pagina or protocolo_pagamento.isdigit():
                                self.driver.implicitly_wait(5)
                                try:
                                    self.driver.find_element(*self.ok).click()
                                except:
                                    pass

                                self.driver.implicitly_wait(30)
                                time.sleep(2)
                                self.driver.find_element(*self.solicitar).click()
                                time.sleep(2)
                                protocolo_encontrado = False

                                while protocolo_encontrado == False:

                                    for i in range(1, 11):

                                        protocolo_portal = self.driver.find_element(By.XPATH, f'/html/body/div[1]/div[5]/section/div/div[10]/div/table/tbody/tr[{i}]/td[4]').text

                                        if protocolo == protocolo_portal or protocolo_pagamento == protocolo_portal:
                                            self.driver.find_element(By.XPATH, f'/html/body/div[1]/div[5]/section/div/div[10]/div/table/tbody/tr[{i}]/td[10]/a[1]').click()
                                            protocolo_encontrado = True
                                            break
                                        
                                    if protocolo_encontrado == False:
                                        self.driver.find_element(*self.proxima_pagina).click()
                                time.sleep(2)

                        else:
                            continue

                    for index, linha in df.iterrows():

                        if f"{linha['Recursado no Portal']}" == "Sim":
                            continue

                        self.driver.implicitly_wait(30)
                        numero_controle = f"{linha['Controle Inicial']}".replace('.0', '')
                        procedimento = f"{linha['Procedimento']}".replace('.0', '')
                        valor_glosado = f"{linha['Valor Glosa']}".replace('.', '')
                        valor_recursar = f"{linha['Valor Recursado']}"
                        justificativa = f"{linha['Recurso Glosa']}"
                        print(f'Controle: {numero_controle}\nProcedimento: {procedimento}\nValor Glosado: {valor_glosado}')
                        self.driver.find_element(*self.guia_input).clear()
                        print('Numero de guia limpado')
                        time.sleep(1)
                        self.driver.find_element(*self.guia_input).send_keys(numero_controle)
                        print('Numero de guia digitado')
                        time.sleep(2)
                        self.driver.find_element(*self.pesquisar).click()
                        print('Guia pesquisada')
                        time.sleep(2)
                        content = self.driver.find_element(*self.div_lista_guias).text

                        if numero_controle not in content:
                            continue

                        self.driver.find_element(*self.guia_click).click()
                        time.sleep(2)
                        table = self.driver.find_element(*self.table).get_attribute('outerHTML')
                        df_tabela = pd.read_html(table)[0]

                        for j in range(1, len(df_tabela) + 1):
                            codigo_portal = self.driver.find_element(By.XPATH, f'/html/body/div[1]/div[5]/section/div/div[7]/div[3]/div[1]/div[2]/fieldset/div/table/tbody/tr[{j}]/td[1]').text
                            valor_glosa_portal = self.driver.find_element(By.XPATH, f'/html/body/div[1]/div[5]/section/div/div[7]/div[3]/div[1]/div[2]/fieldset/div/table/tbody/tr[{j}]/td[5]').text
                            span_class = self.driver.find_element(By.XPATH, f'/html/body/div[1]/div[5]/section/div/div[7]/div[3]/div[1]/div[2]/fieldset/div/table/tbody/tr[{j}]/td[7]/span').get_attribute("class")

                            if procedimento == codigo_portal and valor_glosado == valor_glosa_portal and 'parcial' in span_class:
                                self.driver.find_element(By.XPATH, f'/html/body/div[1]/div[5]/section/div/div[7]/div[3]/div[1]/div[2]/fieldset/div/table/tbody/tr[{j}]/td[7]/a[2]').click()
                                print('Entrou na guia')
                                time.sleep(2)
                                for tentativa in range(0,10):
                                    try:
                                        self.driver.find_element(*self.valor_recursando_input).clear()
                                        time.sleep(2)
                                        print('valor limpado')
                                        break
                                    except:
                                        time.sleep(2)
                                        print('Erro ao limpar valor')
                                time.sleep(2)
                                self.driver.find_element(*self.valor_recursando_input).send_keys(valor_recursar)
                                print('Valor injetado')
                                time.sleep(2)
                                nome_classe = 'panel-title'

                                for number in range(0, 10):
                                    try:
                                        elementos = self.driver.find_elements(By.CLASS_NAME, nome_classe)
                                        print('justificativas encontradas')
                                        break
                                    except:
                                        print('Não encontrou o(s) elemento(s) da justificativa')
                                        time.sleep(2)

                                qtd_elementos = len(elementos)

                                for i in range(1, qtd_elementos + 1):

                                    if qtd_elementos == 1:

                                        for k in range(0, 10):
                                            try:
                                                self.driver.find_element(By.XPATH, '/html/body/div[1]/div[5]/section/div/div[3]/div/div/div[2]/div[1]/div[6]/div[2]/div/div/div[1]/h4/a').click()
                                                print('Justificativa clicada')
                                                time.sleep(2)
                                                break
                                            except:
                                                print('Erro justificativa')
                                                time.sleep(2)

                                    else:
                                        for k in range(0, 10):
                                            try:
                                                self.driver.find_element(By.XPATH, f'/html/body/div[1]/div[5]/section/div/div[3]/div/div/div[2]/div[1]/div[6]/div[2]/div/div[{i}]/div[1]/h4/a').click()
                                                print('Justificativa clicada')
                                                time.sleep(2)
                                                break
                                            except:
                                                print('Erro justificativa')
                                                time.sleep(2)

                                    time.sleep(2)
                                    self.driver.find_element(*self.textarea_justificativa).send_keys(justificativa)
                                    print('Justificativa lançada')
                                    time.sleep(2)
                                    content = self.driver.find_element(*self.xpath_corpo_da_pagina).text
                                    contador = 0

                                    while 'GLOSA TISS' in content:
                                        self.driver.find_element(*self.fechar_justificativa).click()
                                        time.sleep(2)
                                        content = self.driver.find_element(*self.xpath_corpo_da_pagina).text
                                        contador += 1

                                        if contador == 15:
                                            break

                                    print('Fechar justificativa clicado')
                                    time.sleep(2)

                                self.driver.find_element(*self.salvar_recurso).click()
                                print('Recurso salvo')
                                time.sleep(2)
                                content = self.driver.find_element(*self.xpath_corpo_da_pagina).text
                                contador = 0

                                while 'Justificar Evento' in content:
                                    self.driver.find_element(*self.fechar_recurso).click()
                                    time.sleep(2)
                                    content = self.driver.find_element(*self.xpath_corpo_da_pagina).text
                                    contador += 1

                                    if contador == 15:
                                        break

                                print('Recurso Fechado')
                                time.sleep(4)
                                dados = {"Recursado no Portal" : ['Sim']}
                                df_dados = pd.DataFrame(dados)
                                book = load_workbook(planilha)
                                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                                writer.book = book
                                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                                if "P" in protocolo:
                                    df_dados.to_excel(writer, 'Recurso', startrow=index + 1, startcol=21, header=False, index=False)

                                else:
                                    df_dados.to_excel(writer, 'Recurso', startrow=index + 1, startcol=20, header=False, index=False)

                                writer.save()
                                print('Salvo na planilha')
                                writer.close()
                                print('-----------------------------------------------------------------------------------------------------')
                                break

                    self.driver.get('https://servicosonline.cassi.com.br/Prestador/RecursoRevisaoPagamento/TISS/DemonstrativoAnaliseContas/Index')

                tkinter.messagebox.showinfo( 'Recurso CASSI' , f"Recurso concluído!" )
                self.driver.quit()
                break

            except Exception as e:
                print(e)
                self.driver.get('https://servicosonline.cassi.com.br/Prestador/RecursoRevisaoPagamento/TISS/DemonstrativoAnaliseContas/Index')
                continue
#---------------------------------------------------------------------------------------------------------------
def recursar_cassi(user, password):
    try:
        pasta = filedialog.askdirectory()
        global url
        url = 'https://servicosonline.cassi.com.br/GASC/v2/Usuario/Login/Prestador'
        chrome_options = Options()
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')

        options = {
        'proxy': {
                'http': f'http://{user}:{password}@10.0.0.230:3128',
                'https': f'http://{user}:{password}@10.0.0.230:3128'
            }
        }
        try:
            servico = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=servico, options=chrome_options, seleniumwire_options=options)
        except:
            driver = webdriver.Chrome(options=chrome_options, seleniumwire_options=options)

        login_page = Login(driver, url)
        login_page.open()

        login_page.exe_login(
            usuario = "00735860000173",
            senha = "amhpdf123"
        )
        caminho(driver, url).exe_caminho()
        Recursar(driver, url).recurso(pasta=pasta)
    
    except Exception as err:
        tkinter.messagebox.showerror("Automação", f"Ocorreu uma exceção não tratada. \n {err.__class__.__name__} - {err}")
        Pidgin.financeiroDemo(f"Ocorreu uma exceção não tratada. \n {err.__class__.__name__} - {err}")