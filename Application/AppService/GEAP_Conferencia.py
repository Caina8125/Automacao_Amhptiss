from tkinter import filedialog
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchWindowException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from seleniumwire import webdriver
from tkinter.messagebox import showinfo, showerror
from openpyxl import load_workbook
import pandas as pd
import time
import os
from page_element import PageElement

class Login(PageElement):
    acessar_portal = (By.XPATH, '/html/body/div[3]/div[3]/div[1]/form/div[1]/div[1]/div/a')
    usuario = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/div/label[1]/div/div[1]/div/input')
    senha = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[1]/div/label[2]/div/div[1]/div[1]/input')
    entrar = (By.XPATH, '/html/body/div[1]/div[2]/div/div/div[1]/div/div/div/div/div[2]/div/div[2]/button/div[2]/div/div')
    fechar = (By.XPATH, '/html/body/div[4]/div[2]/div/div[3]/button')

    def exe_login(self, senha, cpf):
        try:
            self.driver.find_element(*self.fechar).click()
        except:
            pass
        self.driver.find_element(*self.acessar_portal).click()
        time.sleep(2)
        self.driver.find_element(*self.usuario).send_keys(cpf)
        time.sleep(2)
        self.driver.find_element(*self.senha).send_keys(senha)
        time.sleep(1)
        self.driver.find_element(*self.senha).send_keys(senha)
        time.sleep(2)
        self.driver.find_element(*self.entrar).click()
    
class Caminho(PageElement):
    versao_anterior = (By.XPATH, '/html/body/div[1]/div/div[1]/aside/div[1]/div[3]/button/span[2]/span')
    alerta = (By.XPATH,' /html/body/div[2]/div/center/a')
    portal_tiss = (By.XPATH, '//*[@id="main"]/div/div/div[2]/div[1]/nav/ul/li[20]/a')
    acompanhar_xml = (By.XPATH, '//*[@id="main"]/div/div/div[2]/div/article/div[6]/div[4]/div[4]/div[4]/div/div[2]/ul/li[2]/a')

    def exe_caminho(self):
        time.sleep(4)
        try:
            self.driver.find_element(*self.alerta).click()
        except:
            print('Alerta não apareceu')
        self.driver.implicitly_wait(15)
        self.driver.find_element(*self.versao_anterior).click()
        time.sleep(2)
        self.driver.switch_to.window(self.driver.window_handles[1])
        time.sleep(1)
        self.driver.get('https://www2.geap.org.br/PRESTADOR/portal-tiss.asp')
        self.driver.find_element(*self.acompanhar_xml).click()
        self.driver.switch_to.window(self.driver.window_handles[-1])

class Conferencia(PageElement):
    numero_envio = (By.XPATH, '//*[@id="NroProtocolo"]')
    listar = (By.XPATH, '//*[@id="MenuOptionReport"]')
    table = (By.XPATH, '//*[@id="objTableDetalhe"]')
    guias = (By.XPATH, '//*[@id="objTableAudit"]/tbody/tr[2]/td[7]/a')

    def pesquisar_envio(self, planilha):
        try:
            self.driver.implicitly_wait(30)
            df = pd.read_excel(planilha)
            count_remessa = 0
            arquivo_na_pasta = os.listdir(r'\\10.0.0.239\faturamento\RAMON FAT 239\GEAP - DIVERSOS\GEAP - ROBÔS FAT\Conferência de Anexos\Respostas2')
            verificada_incompleta = False

            if 'Verificadas_Incompleta.xlsx' in arquivo_na_pasta:
                df_processo = pd.read_excel(r'\\10.0.0.239\faturamento\RAMON FAT 239\GEAP - DIVERSOS\GEAP - ROBÔS FAT\Conferência de Anexos\Respostas2\Verificadas_Incompleta.xlsx')
                os.remove(r'\\10.0.0.239\faturamento\RAMON FAT 239\GEAP - DIVERSOS\GEAP - ROBÔS FAT\Conferência de Anexos\Respostas2\Verificadas_Incompleta.xlsx')
                verificada_incompleta = True

            cabecalho = ["Nº Fatura", "Protocolo", "Id Guia", "Guia Prestador", "Arquivo", "Verificação"]
            df_geral = pd.DataFrame(columns=cabecalho)

            for index, linha in df.iterrows():
                count_remessa = count_remessa + 1
                if f"{linha['Observações']}" == "Pesquisado" or f"{linha['Observações']}" == "Erro ao pesquisar número de envio":
                    continue

                numero_processo = (f"{linha['Nº Fatura']}").replace(".0", "")
                numero_envio = (f"{linha['Protocolo']}").replace(".0", "")
                self.driver.find_element(*self.numero_envio).send_keys(numero_envio)
                time.sleep(2)
                self.driver.find_element(*self.listar).click()
                body = self.driver.find_element(By.XPATH, '/html/body').text

                if "Não existem registros na base da dados para o critério escolhido." in body :
                    dado = {'Observações': ["Erro ao pesquisar número de envio"]}
                    df_remessa = pd.DataFrame(dado)
                    book = load_workbook(planilha) #alterar para o endereço do arquivo
                    writer = pd.ExcelWriter(planilha, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df_remessa.to_excel(writer, 'Carta Remessa', startrow= count_remessa, startcol=2, header=False, index=False)
                    writer.save()
                    self.driver.get('https://www2.geap.com.br/PRESTADOR/auditoriadigital/rpt/AcompanhamentoEntrega.aspx')
                    continue
                
                self.driver.find_element(*self.guias).click()
                time.sleep(2)
                self.driver.switch_to.window(self.driver.window_handles[-1])
                count = 0

                if verificada_incompleta == False:
                    table = self.driver.find_element(By.XPATH, '//*[@id="objTableDetalhe"]')
                    tabela_html = table.get_attribute('outerHTML')
                    df_processo = pd.read_html(tabela_html, skiprows=2, header=0)[0]
                    df_processo = df_processo.iloc[:-1]
                    df_processo["Arquivo"] = ""
                    df_processo["Verificação"] = ""
                    df_processo["Nº Fatura"] = numero_processo
                    df_processo["Protocolo"] = numero_envio
                    df_processo = df_processo.loc[:, ["Nº Fatura", "Protocolo", "Id Guia", "Guia Prestador", "Arquivo", "Verificação"]]

                for index2, linha2 in df_processo.iterrows():
                    if linha2["Verificação"] == "Ok" or linha2["Verificação"] == "VERIFICAR" or linha2["Verificação"] == "NÃO ANEXADO":
                        count = count + 1
                        continue
                    time.sleep(2)
                    count = count + 1
                    id_guia = f"{linha2['Id Guia']}".replace(".0", "")
                    self.driver.find_element(By.LINK_TEXT, id_guia).click()
                    try:
                        self.driver.implicitly_wait(5)
                        arquivo = self.driver.find_element(By.XPATH, '/html/body/form/div[3]/table/tbody/tr[6]/td/div/table/tbody/tr[2]/td[2]').text
                        time.sleep(2)
                        numero_fatura = f"{linha2['Guia Prestador']}".replace(".0", "")
                        if numero_fatura in arquivo:
                            contem = "Ok"
                        else:
                            contem = "VERIFICAR"
                        df_processo.loc[index2, 'Arquivo'] = arquivo
                        df_processo.loc[index2, 'Verificação'] = contem
                        self.driver.back()
                        continue
                        
                    except:
                        contem = "NÃO ANEXADO"
                        arquivo = "Não há arquivo"
                        df_processo.loc[index2, 'Arquivo'] = arquivo
                        df_processo.loc[index2, 'Verificação'] = contem
                        self.driver.back()
                        continue
                
                df_geral = pd.concat([df_geral, df_processo], ignore_index=True)
                dado = {'Observações': ["Pesquisado"]}
                df_remessa = pd.DataFrame(dado)
                book = load_workbook(planilha) #alterar para o endereço do arquivo
                writer = pd.ExcelWriter(planilha, engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df_remessa.to_excel(writer, 'Carta Remessa', startrow= count_remessa, startcol=2, header=False, index=False)
                writer.save()
                self.driver.close()
                self.driver.switch_to.window(self.driver.window_handles[-1])
                self.driver.back()
                self.driver.find_element(*self.numero_envio).clear()
                verificada_incompleta = False

            if self.df_existe(df_geral):
                df_geral.to_excel(r"\\10.0.0.239\faturamento\RAMON FAT 239\GEAP - DIVERSOS\GEAP - ROBÔS FAT\Conferência de Anexos\Respostas2\Verificadas.xlsx", index=False)
            else:
                print('A planilha não foi gerada.')
            
        except Exception as e:
            df_geral = pd.concat([df_geral, df_processo], ignore_index=True)

            if self.df_existe(df_geral):
                df_geral.to_excel(r"\\10.0.0.239\faturamento\RAMON FAT 239\GEAP - DIVERSOS\GEAP - ROBÔS FAT\Conferência de Anexos\Respostas2\Verificadas_Incompleta.xlsx", index=False)
            else:
                print('A planilha não foi gerada.')

    def df_existe(self, df):
        try:
            if df.empty == False:
                return True
            else:
                return False
        except:
            return False
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def conferencia(user, password):
    try:
        chrome_options = Options()

        options = {
                'proxy' : {
                    'http': f'http://{user}:{password}@10.0.0.230:3128',
                    'https': f'http://{user}:{password}@10.0.0.230:3128'
                }
            }

        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_argument('--ignore-ssl-errors')



        planilha = filedialog.askopenfilename()
        print(planilha)

        url = "https://www2.geap.com.br/auth/prestador.asp"

        try:
            servico = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=servico, seleniumwire_options=options, options=chrome_options)
            
        except:
            driver = webdriver.Chrome(seleniumwire_options=options, options=chrome_options)

        login_page = Login(driver, url)
        login_page.open()

        login_page.exe_login(
            senha = "Amhp2023",
            cpf = "66661692120"
        )

        time.sleep(2)

        Caminho(driver, url).exe_caminho()

        time.sleep(2)

        Conferencia(driver, url).pesquisar_envio(planilha)

        showinfo( 'Automação' , f"Arquivos anexados!" )

    except NoSuchWindowException as err:
        showerror('Automação', f'A janela do navegador foi fechada!')
    
    except FileNotFoundError as err:
        showerror('Automação', f'Nenhuma planilha foi selecionada!')
    
    except Exception as err:
        showerror("Automação", f"Ocorreu uma exceção não tratada. \n {err.__class__.__name__} - {err}")